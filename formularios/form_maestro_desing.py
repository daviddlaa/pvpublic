import tkinter as tk
from tkinter import font
from config import COLOR_BARRA_SUPERIOR, COLOR_CUERPO_PRINCIPAL, COLOR_MENU_CURSOR_ENCIMA, COLOR_MENU_LATERAL
import util.util_ventana as util_ventana
import util.util_imagenes as util_img
from tkinter import ttk
import sqlite3
import uuid
from customtkinter import *
from tkinter import messagebox
import openpyxl
from datetime import datetime, timedelta
from tkinter import scrolledtext
import random
import string
from tkcalendar import Calendar
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class FormularioMaestroDesing(tk.Tk):
    
    def __init__(self):
        super().__init__()
        self.logo = util_img.leer_imagen("./imagenes/logo.png",(560,136))
        self.perfil = util_img.leer_imagen("./imagenes/perfil.png",(100,100))
        self.menu_burger = util_img.leer_imagen("./imagenes/menu-burger.png",(20,20))
        self.config_window()
        self.paneles()
        self.controles_barra_superior()
        self.controles_menu_lateral()
        self.controles_cuerpo()
        self.editando_producto_id = None
       
    def config_window(self):
        self.title('Punto de Venta')
        self.iconbitmap("./imagenes/logo.ico")
        w, h = 1024, 600
        util_ventana.centrar_ventana(self,w,h)    

    def paneles(self):
        self.barra_superior = tk.Frame(
            self, bg=COLOR_BARRA_SUPERIOR,height=50)
        self.barra_superior.pack(side=tk.TOP,fill="both")

        self.menu_lateral = tk.Frame(self,bg=COLOR_MENU_LATERAL,width=150)
        self.menu_lateral.pack(side=tk.LEFT, fill="both", expand=False)

        self.cuerpo_principal = tk.Frame(
            self, bg=COLOR_CUERPO_PRINCIPAL,width=150)
        self.cuerpo_principal.pack(side=tk.RIGHT, fill="both",expand=True)
        
    def controles_barra_superior(self):

        self.labelTitulo =  tk.Label(self.barra_superior,text="Menu Principal",font=("OCR A Extended", 14))
        
        self.labelTitulo.config(fg="#fff",font=("OCR A Extended", 14), bg=COLOR_BARRA_SUPERIOR,
        pady=10, width=16)
        self.labelTitulo.pack(side=tk.LEFT)

        # Botón del menú lateral
        self.buttonMenuLateral = CTkButton(self.barra_superior, text='\uf0c9',
                                           command=self.toggle_panel,fg_color=COLOR_BARRA_SUPERIOR,font=("OCR A Extended", 14),hover_color=COLOR_BARRA_SUPERIOR)
        self.buttonMenuLateral.pack(side=tk.LEFT)


        self.labelTitulo = tk.Label(
            self.barra_superior,text="@bacosoluciones")
        self.labelTitulo.config(
            fg="#fff",font=("OCR A Extended", 14),bg=COLOR_BARRA_SUPERIOR,padx=10,width=20)
        self.labelTitulo.pack(side=tk.RIGHT)

    def controles_menu_lateral(self):
        ancho_menu = 24
        alto_menu = 2
        font_awesome = font.Font(family='OCR A Extended',size=14)

        self.labelPerfil =tk.Label(
            self.menu_lateral, image=self.perfil, bg=COLOR_MENU_LATERAL)
        self.labelPerfil.pack(side=tk.TOP,pady=10)

        self.buttonInventario = tk.Button(self.menu_lateral,command=self.Inventario) 
        self.buttonIngresoVentas = tk.Button(self.menu_lateral,command=self.ventas) 
        self.buttonHistorialVentas = tk.Button(self.menu_lateral,command=self.historial_ventas_calendario) 
       # self.buttonClientes = tk.Button(self.menu_lateral,command=self.mostrar_clientes)  
       # self.buttonDatosNegocio = tk.Button(self.menu_lateral,command=self.datos_negocio)
        self.buttonUsuarios = tk.Button(self.menu_lateral)
        self.buttonSalir= tk.Button(self.menu_lateral) 
     
        buttons_info = [
        ("INVENTARIO", "\uf494", self.buttonInventario), 
        ("CAJA REGISTRADORA", "\uf788", self.buttonIngresoVentas), 
        ("HISTORIAL VENTAS", "\uf073", self.buttonHistorialVentas), 
        #("CLIENTES", "\uf007", self.buttonClientes),
        #("DATOS NEGOCIO", "\uf54e", self.buttonDatosNegocio),
        ("USUARIOS", "\ue594", self.buttonUsuarios),
        ("SALIR", "\uf2f5", self.buttonSalir),
        ]

        for text, icon, button in buttons_info:
            self.configurar_boton_menu(button, text, icon, font_awesome, ancho_menu, alto_menu)
    
    def toggle_panel(self):
        # Alternar visibilidad del menú lateral
        if self.menu_lateral.winfo_ismapped():
             self.menu_lateral.pack_forget()
           
        else:
            self.menu_lateral.pack(side=tk.LEFT, fill='y')

    def bind_hover_events(self,button):
        button.bind("<Enter>", lambda event: self.on_enter(event,button))
        button.bind("<Leave>", lambda event: self.on_leave(event,button))
    
    def on_enter(self,event,button):
        button.config(bg=COLOR_MENU_CURSOR_ENCIMA,fg='white')
    
    def on_leave(self,event,button):
        button.config(bg=COLOR_MENU_LATERAL,fg='white')
    
    def configurar_boton_menu(self, button, text, icon, font_awesome, ancho_menu, alto_menu): 
        button.config(text=f" {icon} {text}", anchor="w", font=font_awesome,
        bd=0, bg=COLOR_MENU_LATERAL, fg="white", width=ancho_menu, height=alto_menu) 
        button.pack(side=tk.TOP) 
        self.bind_hover_events(button)

    def controles_cuerpo(self):
        label = tk.Label(self.cuerpo_principal,image=self.logo,
                         bg=COLOR_CUERPO_PRINCIPAL)
        label.place(x=0,y=0,relwidth=1,relheight=1)
    

#---------------------funciones para el inventario-------------------------------------------------
    
    def Inventario(self):
       
        # Limpiar cualquier widget existente en el cuerpo principal
        for widget in self.cuerpo_principal.winfo_children():
            widget.destroy()
          # Definir el estilo general para los botones
        
        ESTILO_CTKBOTONES_DATOS_INVENTARIO = {
            'width': 50,
            'height': 40,
            'text_color': 'black',
            'font': ("OCR A Extended", 13)
        }
        ESTILO_TITULO_LABEL_DATOS_INVENTARIO = {
            
            'font': ("OCR A Extended", 14),  
        }
        ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO = {
            'text_color': 'black',
            'font': ("OCR A Extended", 14),
               
        }
        ALINEACION_FORMULARIO = {
        'padx': 5,
        'pady': 5,
        'sticky': 'w'
        }
    
        # Crear un formulario para agregar nuevos elementos al inventario
        formulario_inventario = tk.LabelFrame(self.cuerpo_principal,text="FORMULARIO DE PRODUCTOS: " ,bg=COLOR_CUERPO_PRINCIPAL,**ESTILO_TITULO_LABEL_DATOS_INVENTARIO)
        formulario_inventario.pack(padx=10, pady=10, fill='x', expand=True,ipadx=10,ipady=10)

        
        # Entry oculto para el ID
        id_producto_entry = tk.Entry(formulario_inventario) 
        id_producto_entry.grid(row=0, column=5)
        id_producto_entry.grid_remove()

        # Nombre
        CTkLabel(formulario_inventario, text="Nombre:",**ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO).grid(row=1, column=0, **ALINEACION_FORMULARIO)
        nombre_producto_entry = CTkEntry(formulario_inventario, **ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO)
        nombre_producto_entry.grid(row=1, column=1, **ALINEACION_FORMULARIO)

        # Descripción
        CTkLabel(formulario_inventario, text="Descripción:",**ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO).grid(row=1, column=2, **ALINEACION_FORMULARIO)
        descripcion_entry = CTkEntry(formulario_inventario, **ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO)
        descripcion_entry.grid(row=1, column=3, **ALINEACION_FORMULARIO)

        # Categoría
        CTkLabel(formulario_inventario, text="Categoría:",**ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO).grid(row=1, column=4, **ALINEACION_FORMULARIO)
        categoria_entry = CTkEntry(formulario_inventario, **ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO)
        categoria_entry.grid(row=1, column=5, **ALINEACION_FORMULARIO)

        # Precio venta
        CTkLabel(formulario_inventario, text="Precio venta:",**ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO).grid(row=2, column=0, **ALINEACION_FORMULARIO)
        precio_venta_entry = CTkEntry(formulario_inventario, **ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO)
        precio_venta_entry.grid(row=2, column=1, **ALINEACION_FORMULARIO)

        # Precio compra
        CTkLabel(formulario_inventario, text="Precio compra:",**ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO).grid(row=2, column=2, **ALINEACION_FORMULARIO)
        precio_compra_entry = CTkEntry(formulario_inventario, **ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO)
        precio_compra_entry.grid(row=2, column=3, **ALINEACION_FORMULARIO)

        # Existencias
        CTkLabel(formulario_inventario, text="Existencias:",**ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO).grid(row=2, column=4, **ALINEACION_FORMULARIO)
        existencias_entry = CTkEntry(formulario_inventario, **ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO)
        existencias_entry.grid(row=2, column=5, **ALINEACION_FORMULARIO)

        # Stock Mínimo
        CTkLabel(formulario_inventario, text="Stock Mín:",**ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO).grid(row=3, column=0, **ALINEACION_FORMULARIO)
        stock_minimo_entry = CTkEntry(formulario_inventario, **ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO)
        stock_minimo_entry.grid(row=3, column=1, **ALINEACION_FORMULARIO)

        # Código de barras
        CTkLabel(formulario_inventario, text="Código de barras:",**ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO).grid(row=3, column=2, **ALINEACION_FORMULARIO)
        codigo_barras_entry = CTkEntry(formulario_inventario, **ESTILO_ENTRYS_LABEL_DATOS_INVENTARIO)
        codigo_barras_entry.grid(row=3, column=3, **ALINEACION_FORMULARIO)

      # Botón para agregar el producto al inventario
        agregar_button = CTkButton(formulario_inventario, text="\uf055 Agregar\nProducto",**ESTILO_CTKBOTONES_DATOS_INVENTARIO,
                                   command=lambda:self.agregar_producto(tree_inventario, nombre_producto_entry, descripcion_entry, categoria_entry, precio_venta_entry, 
                         precio_compra_entry, existencias_entry, stock_minimo_entry, codigo_barras_entry, id_producto_entry,descripcion_pro_inventario))
        agregar_button.grid(row=1, column=6,**ALINEACION_FORMULARIO)

       
        limpiar_formulario_button = CTkButton(formulario_inventario, text="\uf1b8 Limpiar\nformulario",**ESTILO_CTKBOTONES_DATOS_INVENTARIO,
                                              command=lambda:self.limpiar_formulario(id_producto_entry,nombre_producto_entry,descripcion_entry,categoria_entry,precio_venta_entry,
                           precio_compra_entry,existencias_entry,stock_minimo_entry,codigo_barras_entry,descripcion_pro_inventario))
        limpiar_formulario_button.grid(row=2, column=6,**ALINEACION_FORMULARIO)
             
        # CREAR EL TREEVIEW -------------------------------------------------
        treeview_inventario_frame = tk.LabelFrame(self.cuerpo_principal,background=COLOR_CUERPO_PRINCIPAL)
        treeview_inventario_frame.pack(padx=10, pady=10, fill='both',ipadx=10,ipady=10)

        # Configurar el Grid para que el frame se expanda con la ventana
        treeview_inventario_frame.grid_rowconfigure(0, weight=1)
        treeview_inventario_frame.grid_columnconfigure(0, weight=1)

        # Configurar el estilo
        style = ttk.Style()
        style.configure("mystyle.Treeview", highlightthickness=0, bd=0, font = ("OCR A Extended", 10))  # Modify the font of the body
        style.configure("mystyle.Treeview.Heading", font = ("OCR A Extended", 9,"bold"))  # Modify the font of the headings
        style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders

        # Crear el TreeView con el estilo configurado y limitar el número máximo de filas
        tree_inventario = ttk.Treeview(treeview_inventario_frame, style="mystyle.Treeview", 
                                        columns=("id","Nombre", "Descripción","Categoria", "Precio Venta", "Precio Compra", "Existencias", "Stock Mínimo", "Código Barras"), 
                                        show="headings")

        tree_inventario.heading("id", text="ID")
        tree_inventario.heading("Nombre", text="NOMBRE")
        tree_inventario.heading("Descripción", text="DESCRIPCION")
        tree_inventario.heading("Categoria", text="CATEGORIA")
        tree_inventario.heading("Precio Venta", text="PVP")
        tree_inventario.heading("Precio Compra", text="P. COMPRA")
        tree_inventario.heading("Existencias", text="EXISTENCIAS")
        tree_inventario.heading("Stock Mínimo", text="STOCK MIN")
        tree_inventario.heading("Código Barras", text="COD BARRAS")

        # Ajustar el ancho de las columnas al cambiar el tamaño de la ventana
        for col in ("Nombre","Descripción","Categoria","Precio Venta", "Precio Compra", "Existencias", "Stock Mínimo", "Código Barras"):
            tree_inventario.column(col, width=60, anchor="center") 
            tree_inventario.column("id", width=0, stretch=tk.NO) 
      
        tree_inventario.grid(row=1, column=0, sticky="nsew")
        tree_inventario.bind('<ButtonRelease-1>', lambda event: self.actualizar_descripcion(event, tree_inventario,descripcion_pro_inventario))
        self.actualizar_treeview(tree_inventario)
         
        scrollbar = CTkScrollbar(treeview_inventario_frame, command=tree_inventario.yview)
        scrollbar.grid(row=1, column=1)
        tree_inventario.configure(yscrollcommand=scrollbar.set)

        entry_buscar = CTkEntry(treeview_inventario_frame,placeholder_text="\uf3eb BUSCAR POR NOMBRE, DESCRIPCION, CATEGORIA O COD BARRAS",width=500,**ESTILO_TITULO_LABEL_DATOS_INVENTARIO)
        entry_buscar.grid(row=0, column=0,pady=(10,0),columnspan=3)
        entry_buscar.bind("<KeyRelease>", lambda event: self.filtrar_productos(event, tree_inventario,entry_buscar))

        descripcion_pro_inventario = scrolledtext.ScrolledText(treeview_inventario_frame, wrap="word", width=20, height=12, font=('OCR A Extended', 13))
        descripcion_pro_inventario.grid(column=2, row=1, padx=(0, 10), pady=10, sticky='w')
        descripcion_pro_inventario.config(bg=COLOR_CUERPO_PRINCIPAL)  

        buttons_frame = tk.LabelFrame(self.cuerpo_principal,text="ACCIONES PARA EL INVENTARIO: ",background=COLOR_CUERPO_PRINCIPAL,**ESTILO_TITULO_LABEL_DATOS_INVENTARIO)
        buttons_frame.pack(side="left",padx=10, pady=10,ipady=10)

        editar_button = CTkButton(buttons_frame, text="\uf044"" Editar",**ESTILO_CTKBOTONES_DATOS_INVENTARIO,command=lambda: self.editar_producto(tree_inventario, 
                        nombre_producto_entry, descripcion_entry, categoria_entry,
                        precio_venta_entry, precio_compra_entry, existencias_entry, stock_minimo_entry, codigo_barras_entry, id_producto_entry))
        editar_button.grid(column=0,row=0,**ALINEACION_FORMULARIO)
        
        eliminar_button = CTkButton(buttons_frame, text="\uf056"" Eliminar", **ESTILO_CTKBOTONES_DATOS_INVENTARIO,command=lambda: self.eliminar_producto(tree_inventario))
        eliminar_button.grid(column=1, row=0,**ALINEACION_FORMULARIO)

        cargar_button = CTkButton(buttons_frame, text="\uf093"" Cargar\nExcel", **ESTILO_CTKBOTONES_DATOS_INVENTARIO,command=lambda: self.cargar_excel(tree_inventario))
        cargar_button.grid(column=2, row=0,**ALINEACION_FORMULARIO)

        descargar_button = CTkButton(buttons_frame, text="\uf019 Respaldo", **ESTILO_CTKBOTONES_DATOS_INVENTARIO,command=self.descargar_productos_excel)
        descargar_button.grid(column=3, row=0,**ALINEACION_FORMULARIO)

        frame_alerta = tk.LabelFrame(self.cuerpo_principal,background=COLOR_CUERPO_PRINCIPAL)
        frame_alerta.pack(side="right",padx=10, pady=10)

          # Crear la etiqueta para la notificación de existencias bajas
        notificacion_existencias_bajas = ParpadeoEtiqueta(frame_alerta,text="") 
        notificacion_existencias_bajas.grid(column=0,row=0,**ALINEACION_FORMULARIO)
        self.verificar_existencias_bajas(tree_inventario, notificacion_existencias_bajas)
        
        informe_stock_min_button = CTkButton(frame_alerta, text="\uf15b"" Informe\nStock\nmínimo", **ESTILO_CTKBOTONES_DATOS_INVENTARIO,command=lambda:self.mostrar_informe_min_stock())
        informe_stock_min_button.grid(column=1,row=0,**ALINEACION_FORMULARIO)

    def verificar_existencias_bajas(self, tree_inventario, notificacion_existencias_bajas):
        try:
            # Conectar a la base de datos
            conn = sqlite3.connect('farmacia.db')
            cursor = conn.cursor()

            # Obtener los datos de productos
            cursor.execute("SELECT existencias, stock_minimo FROM productos")
            productos = cursor.fetchall()

            # Verificar existencias bajas
            existencias_bajas = [producto[0] < producto[1] for producto in productos]

            # Si hay existencias bajas, mostrar notificación
            if any(existencias_bajas):
                notificacion_existencias_bajas.configure(text="¡EXISTEN PRODUCTOS CON EXISTENCIAS BAJAS!\nREVISAR EL INFORME.")
            else:
                notificacion_existencias_bajas.configure(text="")  # Limpiar la notificación si no hay existencias bajas

        except sqlite3.Error as e:
            # Manejar errores de la base de datos
            messagebox.showerror("Error", f"Error de base de datos: {str(e)}")
        except Exception as e:
            # Manejar otros errores
            messagebox.showerror("Error", f"Ha ocurrido un error: {str(e)}")
        finally:
            # Cerrar la conexión a la base de datos
            if conn:
                conn.close()
                
    def actualizar_treeview(self, tree_inventario):
        try:
            # Conectar a la base de datos
            conn = sqlite3.connect('farmacia.db')
            cursor = conn.cursor()

            # Obtener los datos actualizados de productos de la base de datos
            cursor.execute("SELECT * FROM productos")
            productos = cursor.fetchall()

            # Limpiar el TreeView antes de volver a cargar los datos
            tree_inventario.delete(*tree_inventario.get_children())

            # Insertar los nuevos datos en el TreeView
            for producto in productos:
                precio_venta = f"${producto[4]}"  # El índice 3 corresponde al precio de venta
                precio_compra = f"${producto[5]}"  # El índice 4 corresponde al precio de compra
                codigo_barras = producto[8]  # El índice 7 corresponde al código de barras
                tree_inventario.insert("", "end", values=(producto[0], producto[1], producto[2], producto[3], precio_venta, precio_compra, producto[6], producto[7], codigo_barras))
       
        except Exception as e:
            # Mostrar mensaje de error
            messagebox.showerror("Error", f"Ha ocurrido un error al actualizar el TreeView: {str(e)}")
        finally:
            # Cerrar la conexión a la base de datos
            if conn:
                conn.close()

    def filtrar_productos(self, event, tree_inventario, entry_buscar):

        # Obtener el texto ingresado en el Entry de búsqueda
        texto_busqueda = entry_buscar.get().lower()  # Convertir a minúsculas para una búsqueda sin distinción entre mayúsculas y minúsculas

        # Limpiar el TreeView antes de aplicar el filtro
        for row in tree_inventario.get_children():
            tree_inventario.delete(row)

        # Obtener los datos de productos de la base de datos
        conn = sqlite3.connect('farmacia.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM productos")
        productos = cursor.fetchall()

        # Filtrar los productos que coinciden con la búsqueda
        for producto in productos:
            # Convertir todos los campos a minúsculas para una búsqueda sin distinción entre mayúsculas y minúsculas
            if any(texto_busqueda in str(valor).lower() for valor in producto):
                precio_venta = f"${producto[4]}"  # El índice 3 corresponde al precio de venta
                precio_compra = f"${producto[5]}"  # El índice 4 corresponde al precio de compra
                tree_inventario.insert("", "end", text="", values=(producto[0],producto[1], producto[2],producto[3], precio_venta, precio_compra, producto[6], producto[7], producto[8]))

        # Cerrar la conexión a la base de datos
        conn.close()

    def actualizar_descripcion(self, event, tree_inventario, descripcion_pro_inventario):
        try:
            # Obtener el índice del item seleccionado en el TreeView
            item_seleccionado = tree_inventario.focus()
            
            # Obtener los valores del item seleccionado
            valores = tree_inventario.item(item_seleccionado, 'values')
            
            # Verificar si hay valores en la fila seleccionada
            if valores:
                # Obtener la descripción del producto
                descripcion_producto = valores[2]  
                # Actualizar el texto del label con la descripción del producto
                descripcion_pro_inventario.config(state=tk.NORMAL)  # Habilitar la edición del ScrolledText
                descripcion_pro_inventario.delete('1.0', tk.END)  # Limpiar el contenido anterior
                descripcion_pro_inventario.insert(tk.END, descripcion_producto)  # Insertar la nueva descripción
                descripcion_pro_inventario.config(state=tk.DISABLED)  # Deshabilitar la edición del ScrolledText
            else:
                # Si no hay valores en la fila seleccionada, mostrar un mensaje de error
                descripcion_pro_inventario.config(state=tk.NORMAL)  # Habilitar la edición del ScrolledText
                descripcion_pro_inventario.delete('1.0', tk.END)  # Limpiar el contenido si no hay producto seleccionado
                descripcion_pro_inventario.insert(tk.END, "Información General: Seleccione un producto para ver su información.")
                #descripcion_pro_inventario.config(state=tk.DISABLED)  # Deshabilitar la edición del ScrolledText
        except Exception as e:
            # Manejar cualquier error y mostrar un mensaje
            messagebox.showerror("Error", f"Error al actualizar la descripción: {str(e)}")

    def eliminar_producto(self, tree_inventario):
        try:
            # Obtener el índice del item seleccionado en el TreeView
            item_seleccionado = tree_inventario.focus()

            if item_seleccionado:
                # Mostrar un cuadro de diálogo de confirmación
                confirmar_eliminar = messagebox.askyesno("Confirmar Eliminación", "¿Está seguro de que desea eliminar este producto?")

                if confirmar_eliminar:
                    # Obtener los valores de la fila seleccionada
                    valores = tree_inventario.item(item_seleccionado, 'values')

                    # Obtener el ID del producto seleccionado (asumiendo que el ID está en la primera columna)
                    id_producto = valores[0]  # Suponiendo que el ID del producto está en la primera columna

                    # Conectar a la base de datos
                    conn = sqlite3.connect('farmacia.db')
                    cursor = conn.cursor()

                    # Eliminar el producto de la tabla de productos
                    cursor.execute("DELETE FROM productos WHERE id=?", (id_producto,))

                    # Confirmar los cambios en la base de datos
                    conn.commit()

                    # Eliminar el producto seleccionado del TreeView
                    tree_inventario.delete(item_seleccionado)

                    # Mostrar mensaje de éxito
                    messagebox.showinfo("Éxito", "El producto se ha eliminado correctamente")
            else:
                # Si no se selecciona ningún elemento, mostrar un mensaje de advertencia
                messagebox.showwarning("Advertencia", "Por favor, seleccione un producto para eliminar")

        except sqlite3.Error as e:
            # Mostrar mensaje de error específico de SQLite
            messagebox.showerror("Error", f"Error de base de datos: {str(e)}")
        except Exception as e:
            # Mostrar mensaje de error genérico
            messagebox.showerror("Error", f"Ha ocurrido un error al eliminar el producto: {str(e)}")
        finally:
            # Cerrar la conexión a la base de datos si está abierta
            if conn:
                conn.close()

    def editar_producto(self, tree_inventario, nombre_producto_entry, descripcion_entry, categoria_entry,
                        precio_venta_entry, precio_compra_entry, existencias_entry, stock_minimo_entry, codigo_barras_entry, id_producto_entry):
        try:
            # Obtener el índice del item seleccionado en el TreeView
            item_seleccionado = tree_inventario.focus()

            if item_seleccionado:
                # Obtener los valores de la fila seleccionada
                valores = tree_inventario.item(item_seleccionado, 'values')

                # Actualizar self.editando_producto_id con el ID del producto que se está editando
                self.editando_producto_id = valores[0]

                # Llenar los campos del formulario con los valores obtenidos
                nombre_producto_entry.delete(0, 'end')
                nombre_producto_entry.insert(0, valores[1])
                descripcion_entry.delete(0, 'end')
                descripcion_entry.insert(0, valores[2])
                categoria_entry.delete(0, 'end')
                categoria_entry.insert(0, valores[3])
                precio_venta_entry.delete(0, 'end')
                precio_venta_entry.insert(0, valores[4])
                precio_compra_entry.delete(0, 'end')
                precio_compra_entry.insert(0, valores[5])
                existencias_entry.delete(0, 'end')
                existencias_entry.insert(0, valores[6])
                stock_minimo_entry.delete(0, 'end')
                stock_minimo_entry.insert(0, valores[7])
                codigo_barras_entry.delete(0, 'end')
                codigo_barras_entry.insert(0, valores[8])

                # Llenar el campo de ID del producto
                id_producto_entry.delete(0, 'end')
                id_producto_entry.insert(0, self.editando_producto_id)
            else:
                # Si no se selecciona ningún elemento, mostrar un mensaje de advertencia
                messagebox.showwarning("Advertencia", "Por favor, seleccione un producto para editar")

        except Exception as e:
            # Mostrar mensaje de error
            messagebox.showerror("Error", f"Ha ocurrido un error al editar el producto: {str(e)}")

    def agregar_producto(self, tree_inventario, nombre_producto_entry, descripcion_entry, categoria_entry, precio_venta_entry, 
                         precio_compra_entry, existencias_entry, stock_minimo_entry, codigo_barras_entry, id_producto_entry,descripcion_pro_inventario):
        try:
            # Obtener los valores de los campos del formulario
            nombre = nombre_producto_entry.get()
            descripcion = descripcion_entry.get()
            categoria = categoria_entry.get()
            precio_venta = precio_venta_entry.get()
            precio_compra = precio_compra_entry.get()
            existencias = existencias_entry.get()
            stock_minimo = stock_minimo_entry.get()
            codigo_barras = codigo_barras_entry.get()
            id_producto = id_producto_entry.get()

            # Validar que no haya campos vacíos
            if not all([nombre, descripcion, categoria, precio_venta, precio_compra, existencias, stock_minimo, codigo_barras]):
                messagebox.showerror("Error", "Todos los campos son obligatorios")
                return

            # Convertir los valores de precio y existencias a números
            precio_venta = float(precio_venta.replace('$', ''))
            precio_compra = float(precio_compra.replace('$', ''))
            existencias = int(existencias)
            stock_minimo = int(stock_minimo)

            # Validar que los precios y existencias sean valores válidos
            if precio_venta <= 0 or precio_compra <= 0 or existencias < 0 or stock_minimo < 0:
                messagebox.showerror("Error", "Los precios y existencias deben ser mayores que cero, y las existencias no pueden ser negativas")
                return

            # Conectar a la base de datos
            conn = sqlite3.connect('farmacia.db')
            cursor = conn.cursor()

            # Verificar si el código de barras ya existe en la base de datos
            cursor.execute("SELECT * FROM productos WHERE codigo_barras=?", (codigo_barras,))
            existing_product = cursor.fetchone()

            # Verificar si estamos editando un producto existente o agregando uno nuevo
            if self.editando_producto_id:  # Si hay un ID de producto en modo de edición
                # Actualizar los datos del producto existente
                cursor.execute("UPDATE productos SET nombre=?, descripcion=?, categoria=?, precio_venta=?, precio_compra=?, existencias=?, stock_minimo=?, codigo_barras=? WHERE id=?",
                                (nombre, descripcion, categoria, precio_venta, precio_compra, existencias, stock_minimo, codigo_barras, self.editando_producto_id))
                messagebox.showinfo("Éxito", "El producto se ha actualizado correctamente")
                conn.commit()
            elif existing_product:
                messagebox.showinfo("Info", "Ya existe un producto con el mismo código de barras en la base de datos")
                return
            else:
                # Generar un id único de 8 caracteres alfanuméricos
                producto_id = str(uuid.uuid4())[:8]  # Obtiene los primeros 8 caracteres del id generado

                # Insertar los datos en la tabla de productos
                cursor.execute("INSERT INTO productos (id, nombre, descripcion, categoria, precio_venta, precio_compra, existencias, stock_minimo, codigo_barras) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                                (producto_id, nombre, descripcion, categoria, precio_venta, precio_compra, existencias, stock_minimo, codigo_barras))
                messagebox.showinfo("Éxito", "El producto se ha agregado correctamente")
                conn.commit()

            # Limpiar el TreeView antes de volver a cargar los datos
            tree_inventario.delete(*tree_inventario.get_children())

            # Obtener los datos actualizados de productos de la base de datos
            cursor.execute("SELECT * FROM productos")
            productos = cursor.fetchall()

            for producto in productos:
                precio_venta = f"${producto[4]}"  # El índice 3 corresponde al precio de venta
                precio_compra = f"${producto[5]}"  # El índice 4 corresponde al precio de compra
                tree_inventario.insert("", "end", values=(producto[0], producto[1], producto[2], producto[3], precio_venta, precio_compra, producto[6], producto[7], producto[8]))

        except ValueError:
            # Mostrar mensaje de error si los campos de precio y existencias no son números válidos
            messagebox.showerror("Error", "Los campos de precio y existencias deben ser números")
        except sqlite3.Error as e:
            # Mostrar mensaje de error específico de SQLite
            messagebox.showerror("Error", f"Error de base de datos: {str(e)}")
        except Exception as e:
            # Mostrar mensaje de error genérico
            messagebox.showerror("Error", f"Ha ocurrido un error al agregar el producto: {str(e)}")
        finally:
            # Cerrar la conexión a la base de datos si está abierta
            if conn:
                conn.close()
                self.editando_producto_id = None  # Restablecer la variable de edición
        self.limpiar_formulario(id_producto_entry,nombre_producto_entry,descripcion_entry,categoria_entry,precio_venta_entry,
                           precio_compra_entry,existencias_entry,stock_minimo_entry,codigo_barras_entry,descripcion_pro_inventario)

    def limpiar_formulario(self,id_producto_entry,nombre_producto_entry,descripcion_entry,categoria_entry,precio_venta_entry,
                           precio_compra_entry,existencias_entry,stock_minimo_entry,codigo_barras_entry,descripcion_pro_inventario):
        
              # Limpiar todos los campos de entrada
        id_producto_entry.delete(0,'end')
        nombre_producto_entry.delete(0, 'end')
        descripcion_entry.delete(0, 'end')
        categoria_entry.delete(0, 'end')
        precio_venta_entry.delete(0, 'end')
        precio_compra_entry.delete(0, 'end')
        existencias_entry.delete(0, 'end')
        stock_minimo_entry.delete(0, 'end')
        codigo_barras_entry.delete(0, 'end')
        descripcion_pro_inventario.delete(1.0, tk.END)
    
    def cargar_excel(self, tree_inventario):
        try:
            # Abrir el cuadro de diálogo para seleccionar un archivo Excel
            filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])

            # Conectar a la base de datos
            conn = sqlite3.connect('farmacia.db')
            cursor = conn.cursor()

            # Leer el archivo Excel directamente con openpyxl
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active

            # Contadores para el informe
            filas_procesadas = 0
            productos_insertados = 0

            # Iterar sobre cada fila del archivo Excel
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Obtener los valores de cada columna
                producto_id = row[0].strip() if row[0] and row[0].strip() else str(uuid.uuid4())[:8]  # Suponiendo que la primera columna es el ID
                nombre = row[1]
                descripcion = row[2]
                categoria = row[3]
                precio_venta = row[4]
                precio_compra = row[5]
                existencias = row[6]
                stock_minimo = row[7]
                codigo_barras = row[8]

                # Insertar los datos en la tabla de productos
                cursor.execute("INSERT INTO productos (id, nombre, descripcion,categoria, precio_venta, precio_compra, existencias, stock_minimo, codigo_barras) VALUES (?, ?, ?, ?, ?, ?, ?, ?,?)",
                                (producto_id, nombre, descripcion,categoria, precio_venta, precio_compra, existencias, stock_minimo, codigo_barras))

                # Incrementar el contador de productos insertados
                productos_insertados += 1

                # Incrementar el contador de filas procesadas
                filas_procesadas += 1

            # Confirmar los cambios en la base de datos
            conn.commit()

            # Mostrar mensaje de éxito con el informe
            messagebox.showinfo("Éxito", f"Se procesaron {filas_procesadas} archivo(s) y se insertaron {productos_insertados} productos correctamente desde el (los) archivo Excel.")

            # Actualizar el TreeView con los nuevos datos
            self.actualizar_treeview(tree_inventario)

        except Exception as e:
            # Mostrar mensaje de error
            messagebox.showerror("Error", f"Ha ocurrido un error al cargar los productos desde el archivo Excel: {str(e)}")
        finally:
            # Cerrar la conexión a la base de datos
            if conn:
                conn.close()

    def descargar_productos_excel(event=None):
        try:
            # Conectar a la base de datos y obtener los datos de productos
            conn = sqlite3.connect('farmacia.db')
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM productos")
            productos = cursor.fetchall()
            
            # Crear un nuevo libro de Excel
            libro_excel = openpyxl.Workbook()
            hoja_excel = libro_excel.active
            hoja_excel.title = "productos"
            
            # Agregar encabezados de columna (asegúrate de que los nombres coincidan con los nombres de las columnas en la base de datos)
            hoja_excel.append(["id", "nombre", "descripcion", "categoria","precio_venta", "precio_compra", "existencias", "stock_minimo", "codigo_barras"])
            
            # Agregar datos de productos a las filas
            for producto in productos:
                hoja_excel.append(producto)
            
            # Crear una carpeta para el backup si no existe
            backup_folder = 'backup_productos'
            if not os.path.exists(backup_folder):
                os.makedirs(backup_folder)
            
            # Generar el nombre del archivo con la fecha y hora del backup
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"productos_{timestamp}.xlsx"
            filepath = os.path.join(backup_folder, filename)
            
            # Guardar el archivo Excel en la carpeta de backup
            libro_excel.save(filepath)
            
            # Mostrar mensaje de éxito
            messagebox.showinfo("Éxito", f"La tabla de productos se ha descargado correctamente como {filename}")
        except Exception as e:
            # Mostrar mensaje de error
            messagebox.showerror("Error", f"Ha ocurrido un error al descargar la tabla de productos: {str(e)}")
        finally:
            # Cerrar la conexión a la base de datos
            conn.close()

    def mostrar_informe_min_stock(self):
        # Verificar si la ventana ya está abierta
        if hasattr(self, "informe_window") and self.informe_window.winfo_exists():
            self.informe_window.deiconify()  # Enfocar ventana existente si está abierta
            self.actualizar_informe_min_stock_desde_bd()  # Actualizar contenido desde la base de datos
            return
            
        # Crear una nueva ventana
        self.informe_window = tk.Toplevel(self.cuerpo_principal, background=COLOR_CUERPO_PRINCIPAL)
        self.informe_window.title("INFORME")
        self.informe_window.iconbitmap("./imagenes/logo.ico")
        self.informe_window.geometry("300x200")
        self.informe_window.resizable(False, False)

        # Crear un Frame para contener el Treeview y el Scrollbar
        tree_frame = tk.Frame(self.informe_window)
        tree_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        # Configurar el estilo
        style = ttk.Style()
        style.configure("mystyle.Treeview", highlightthickness=0, bd=0, font=("OCR A Extended", 10))  # Modify the font of the body
        style.configure("mystyle.Treeview.Heading", font=("OCR A Extended", 9, "bold"))  # Modify the font of the headings
        style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders

        # Crear un Treeview en el frame con el estilo configurado
        self.informe_treeview = ttk.Treeview(tree_frame, style="mystyle.Treeview")
        self.informe_treeview.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Crear un Scrollbar para el Treeview
        treeview_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.informe_treeview.yview)
        treeview_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Configurar el comando de desplazamiento del Treeview
        self.informe_treeview.configure(yscrollcommand=treeview_scrollbar.set)

        # Configurar las columnas del Treeview
        self.informe_treeview["columns"] = ("Nombre", "Existencias")
        self.informe_treeview.heading("#0", text="ID", anchor="center")  # Oculta la columna de índice
        self.informe_treeview.heading("Nombre", text="Nombre")
        self.informe_treeview.heading("Existencias", text="Existencias")

        # Ajustar el ancho de las columnas
        self.informe_treeview.column("#0", width=0, stretch=tk.NO)  # Oculta la columna de índice
        self.informe_treeview.column("Nombre", width=150, anchor="center")
        self.informe_treeview.column("Existencias", width=100, anchor="center")
        
        # Obtener y mostrar los productos con existencias mínimas desde la base de datos
        self.actualizar_informe_min_stock_desde_bd()

    def actualizar_informe_min_stock_desde_bd(self):
        # Limpiar contenido anterior
        for item in self.informe_treeview.get_children():
            self.informe_treeview.delete(item)

        # Conectar a la base de datos
        conexion = sqlite3.connect('farmacia.db')
        cursor = conexion.cursor()

        # Realizar la consulta SQL para seleccionar productos con existencias por debajo del stock mínimo
        cursor.execute("SELECT id, nombre, existencias, stock_minimo FROM productos WHERE existencias < stock_minimo")
        productos_con_existencias_bajas = cursor.fetchall()

        # Mostrar los productos en el Treeview
        for producto in productos_con_existencias_bajas:
            self.informe_treeview.insert("", "end", values=(producto[1], producto[2]))

        # Cerrar la conexión con la base de datos
        conexion.close()

#---------------------FUNCIONES PARA VENTAS--------------------------------

    def ventas(self):
        # Limpiar cualquier widget existente en el cuerpo principal
        for widget in self.cuerpo_principal.winfo_children():
            widget.destroy()

#-----------------------FRAMES -----------------------------------------------------------------------------------
        botones_accesos_rapidos = tk.LabelFrame(self.cuerpo_principal, background=COLOR_CUERPO_PRINCIPAL)
        botones_accesos_rapidos.pack(fill='x', padx=10, pady=(10, 0))
        

        filtrado_productos_venta = tk.LabelFrame(self.cuerpo_principal, text="BUSQUEDA PRODUCTOS", 
                                                 background=COLOR_CUERPO_PRINCIPAL, font=("OCR A Extended",12))
        filtrado_productos_venta.pack(fill='x', padx=10, pady=10)
       

        detalles_venta = tk.LabelFrame(self.cuerpo_principal, text="DETALLES", 
                                       background=COLOR_CUERPO_PRINCIPAL, font=("OCR A Extended",12))
        detalles_venta.pack(fill="x",side="left", padx=10, pady=(10, 0))
       

        detalles_acciones = tk.LabelFrame(self.cuerpo_principal, 
                                       background=COLOR_CUERPO_PRINCIPAL, font=("OCR A Extended",12))
        detalles_acciones.pack(fill="none",side="left", padx=10, pady=(10, 0))
       

        detalles_totales_frame = tk.LabelFrame(self.cuerpo_principal,text="TOTALES",
                                               background=COLOR_CUERPO_PRINCIPAL, font=("OCR A Extended",12))
        detalles_totales_frame.pack(fill="none",side="left", padx=10, pady=(10, 0))
       

        # Crear los botones de acceso rápido
        boton_nueva_venta = CTkButton(botones_accesos_rapidos, text="NUEVA\nVENTA",width=70,height=70,text_color='black',font=("OCR A Extended",12))
        boton_nueva_venta.grid(column=0, row=0, padx=10, pady=10)
        boton_nueva_venta.configure(command=lambda: self.restablecer_valores(entry_busqueda_producto, treeview, lbl_pro_sel, lbl_pro_descripcion,
                                                                     entry_precio_venta, entry_cantidad, lbl_subtotal_venta,
                                                                     detalles_treeview,entry_nombre_producto_noinv, entry_precio_producto_detalle, 
                                                                     entry_cantidad_producto_detalle, lbl_total_venta,entry_valor_recibido,lbl_vuelto))

        inicio_caja = CTkButton(botones_accesos_rapidos, text="INICIO\nDE\nOPERACIONES",width=70, height=70, text_color='black',font=("OCR A Extended",12),
                                command=lambda: self.formulario_inicio_operaciones())
        inicio_caja.grid(column=2,row=0, padx=10, pady=10)

        cierre_caja = CTkButton(botones_accesos_rapidos, text="CIERRE\nDE\nCAJA",width=70, height=70, text_color='black',font=("OCR A Extended",12),
                                command=lambda:self.formulario_cierre_caja())
        cierre_caja.grid(column=3,row=0, padx=10, pady=10)
        
        boton_movimientos_caja = CTkButton(botones_accesos_rapidos,text="MOVIMIENTOS\nDE\nCAJA",width=70, height=70, 
                                           text_color='black',font=("OCR A Extended",12),command=lambda: self.formulario_movimientos_caja())
        boton_movimientos_caja.grid(column=4,row=0, padx=10, pady=10)

        etiqueta_estado_operaciones = ParpadeoEtiqueta(botones_accesos_rapidos,text="JORNADA EN CURSO",font=("OCR A Extended",16))
        etiqueta_estado_operaciones.grid(column=5,row=0, padx=10, pady=10)

        # Entry para ingresar el término de búsqueda
        entry_busqueda_producto = CTkEntry(filtrado_productos_venta, bg_color=COLOR_CUERPO_PRINCIPAL, width=300, 
                                        placeholder_text="\uf02a Ingrese Cod barras, Nombre o Descripcion",font=("OCR A Extended",14))
        entry_busqueda_producto.grid(column=0, row=0, padx=(10, 5), pady=10, sticky='ew')
        # Vincular la función de actualización al evento de modificación del Entry
        entry_busqueda_producto.bind('<KeyRelease>', lambda event: self.actualizar_filtrado(event, entry_busqueda_producto, treeview, lbl_pro_sel,lbl_pro_descripcion))

        # Crear el estilo para el Treeview
        style = ttk.Style()
        style.configure("mystyle.Treeview", highlightthickness=0, bd=0, font=('OCR A Extended', 10))
        style.configure("mystyle.Treeview.Heading", font=('OCR A Extended', 11, 'bold'))
        

        # Aplicar el estilo al Treeview
        treeview = ttk.Treeview(filtrado_productos_venta, columns=('id','Nombre', 
                                                                   'Descripción', 
                                                                   'Precio',
                                                                   'Existencias',
                                                                   'Codigo Barras'), height=5, style="mystyle.Treeview")
        treeview.grid(column=0, row=1, columnspan=1, padx=10, pady=10, sticky='nsew')
        
        # Ocultar las columnas
        for col in ['#0','#1', '#3', '#4', '#6']:
           treeview.column(col, width=0, stretch=tk.NO)
        for col in  ['#2','#5']:
           treeview.column(col, anchor="center")
       

        # Asignar los encabezados
        treeview.heading('#0', text="")
        treeview.heading('#1', text="id")
        treeview.heading('#2', text='Nombre')
        treeview.heading('#3', text='Descripción')
        treeview.heading('#4', text='Precio')
        treeview.heading('#5', text='Existencias')
        treeview.heading('#6', text='Codigo Barras')

        treeview.grid(column=0, row=1, columnspan=1)

        treeview.bind('<KeyPress>', lambda event: self.move_selection_filtrado(event, treeview,entry_cantidad))

        # Vincular la función de actualización del precio al evento de selección en el Treeview
        treeview.bind("<<TreeviewSelect>>", lambda event: self.actualizar_precio_venta(event, entry_precio_venta, treeview, 
                                                                                       entry_cantidad, lbl_subtotal_venta, lbl_pro_sel, descripcion, lbl_pro_descripcion,entry_id_producto))

       # Fondo deseado
        fondo = "#F9F9FA"

        # Crear los widgets ScrolledText con el fondo deseado
        lbl_pro_sel = scrolledtext.ScrolledText(filtrado_productos_venta, wrap="word", width=20, height=2, font=('OCR A Extended', 13))
        lbl_pro_sel.grid(column=2, row=0, padx=(0, 10), pady=10, sticky='w')
        lbl_pro_sel.config(bg=fondo)  # Establecer el fondo

        lbl_pro_descripcion = scrolledtext.ScrolledText(filtrado_productos_venta, wrap="word", width=20, height=7, font=('OCR A Extended', 13))
        lbl_pro_descripcion.grid(column=2, row=1, columnspan=2, padx=(0, 10), pady=10, sticky='w')
        lbl_pro_descripcion.config(bg=fondo)  # Establecer el fondo

        separador1 = ttk.Separator(filtrado_productos_venta, orient="vertical")
        separador1.grid(column=3, row=0,rowspan=3, sticky="ns",padx=5,pady=5)
        
        entry_id_producto=tk.Entry(filtrado_productos_venta)
        entry_id_producto.grid_forget
        
        entry_nombre_producto_noinv = CTkEntry(filtrado_productos_venta, placeholder_text="INGRESE NOMBRE PRODUCTO",font=("OCR A Extended", 12),width=230)
        entry_nombre_producto_noinv.grid(column=4, row=0,columnspan=2)
        entry_nombre_producto_noinv.bind("<Button-1>",lambda event: self.deselect_item(event,lbl_pro_sel,lbl_pro_descripcion,treeview,
                                                                                       entry_precio_venta,entry_cantidad,entry_busqueda_producto,lbl_subtotal_venta))
        
        descripcion = 0

        entry_precio_venta = CTkEntry(filtrado_productos_venta, placeholder_text="PRECIO $", width=90, height=70, font=("OCR A Extended", 20))
        entry_precio_venta.grid(column=4, row=1)#, sticky="w", padx=5, pady=5)
        entry_precio_venta.bind('<KeyRelease>', lambda event: self.actualizar_precio_cantidad(event, entry_precio_venta, entry_cantidad, lbl_subtotal_venta))

        entry_cantidad = CTkEntry(filtrado_productos_venta, placeholder_text="\ue43c", width=90, height=70, font=("OCR A Extended", 20))
        entry_cantidad.grid(column=5, row=1)#, sticky='W', padx=5, pady=5)

        entry_cantidad.bind('<KeyRelease>', lambda event: self.actualizar_precio_cantidad(event, entry_precio_venta, entry_cantidad, lbl_subtotal_venta))
        entry_cantidad.bind("<Return>", lambda event: self.agregar_articulo(treeview, entry_cantidad, entry_precio_venta, entry_nombre_producto_noinv, detalles_treeview, lbl_total_venta))
        entry_cantidad.bind("<space>", lambda event: self.agregar_articulo(treeview, entry_cantidad, entry_precio_venta, entry_nombre_producto_noinv, detalles_treeview, lbl_total_venta))

        lbl_subtotal_venta = CTkLabel(filtrado_productos_venta, text="subtotal\n$", width=70, height=50, font=("OCR A Extended", 20), anchor="center", bg_color="#F9F9FA")
        lbl_subtotal_venta.grid(column=6, row=1,sticky='nwse', padx=10, pady=10)

        
        btn_agregar_articulo = CTkButton(filtrado_productos_venta, text="\uf217 Agregar\narticulo", width=70, height=70, font=("OCR A Extended", 14), 
                                         text_color='black', command=lambda: self.agregar_articulo(treeview, entry_cantidad, entry_precio_venta, entry_nombre_producto_noinv, detalles_treeview, lbl_total_venta))
        btn_agregar_articulo.grid(column=6, row=0,sticky='nwse', padx=10, pady=10)

        # Actualizar el filtrado para mostrar todos los productos inicialmente
        self.actualizar_filtrado(None, entry_busqueda_producto, treeview, lbl_pro_sel,lbl_pro_descripcion,entry_id_producto)
        # Llamar a actualizar_precio_venta al final de la función ventas
        self.actualizar_precio_venta(None, entry_precio_venta, treeview, entry_cantidad, lbl_subtotal_venta, lbl_pro_sel,entry_id_producto, None, None)

        # Crear el estilo para el Treeview
        style = ttk.Style()
        style.configure("mystyle.Treeview", highlightthickness=0, bd=0, font=('OCR A Extended', 10), background="#F9F9FA")
        style.configure("mystyle.Treeview.Heading", font=('OCR A Extended', 11, 'bold'), background="#F9F9FA")

        # Aplicar el estilo al Treeview
        detalles_treeview = ttk.Treeview(detalles_venta, columns=("id","Producto", "Cantidad", "Precio", "SubTotal"), 
                                        height=9, style="mystyle.Treeview")

        # Ocultar la primera columna
        detalles_treeview.column("#0", width=0, stretch=tk.NO)
        detalles_treeview.column("#1", width=0,stretch=tk.NO)
        # Configurar las columnas del Treeview
        detalles_treeview.heading("#0", text="")
        detalles_treeview.heading("#1", text="ID")
        detalles_treeview.heading("#2", text="Producto")
        detalles_treeview.heading("#3", text="Precio")
        detalles_treeview.heading("#4", text="Cantidad")
        detalles_treeview.heading("#5", text="Subtotal")

        # Ajustar el ancho de las columnas
        detalles_treeview.column("Producto", width=130, anchor="center")
        detalles_treeview.column("Cantidad", width=90, anchor="center")
        detalles_treeview.column("Precio", width=90, anchor="center")
        detalles_treeview.column("SubTotal", width=90, anchor="center")

        detalles_treeview.grid(column=0, row=0, columnspan=1, padx=10, pady=10, sticky='nsew')
        id_det=0
        detalles_treeview.bind('<KeyPress>', lambda event: self.move_selection_detalles(event, detalles_treeview))
        detalles_treeview.bind("<<TreeviewSelect>>", lambda event: self.on_treeview_select(event, detalles_treeview, id_det, entry_id_producto_detalle, entry_precio_producto_detalle, entry_cantidad_producto_detalle, lbl_pro_sel, lbl_pro_descripcion, treeview, entry_precio_venta, entry_cantidad, entry_busqueda_producto, lbl_subtotal_venta))

        lbl_precio_producto_detalle =CTkLabel(detalles_acciones,text="PRECIO",font=("OCR A Extended",15))
        lbl_precio_producto_detalle.grid(column=2,row=0,padx=10,pady=10)

        entry_id_producto_detalle = tk.Entry(detalles_acciones)
        entry_id_producto_detalle.grid_forget

        entry_cantidad_producto_detalle = CTkEntry(detalles_acciones,placeholder_text="\uf53d", 
                                                 width=70,height=70,font=("OCR A Extended",20))
        entry_cantidad_producto_detalle .grid(column=2, row=1, padx=10, pady=10)

        lbl_cantidad_producto_detalle = CTkLabel(detalles_acciones,text="CANTIDAD",font=("OCR A Extended",15))
        lbl_cantidad_producto_detalle.grid(column=3, row=0, padx=10, pady=10)


        entry_precio_producto_detalle = CTkEntry(detalles_acciones,placeholder_text="\ue43c", 
                                                   width=70,height=70,font=("OCR A Extended",20))                                              
        entry_precio_producto_detalle.grid(column=3, row=1, padx=10, pady=10)

        
        boton_editar_detalle_venta = CTkButton(detalles_acciones, text="REALIZAR\nCAMBIO", 
                                       width=70, height=50, text_color='black', 
                                       font=("OCR A Extended", 12))
        boton_editar_detalle_venta.grid(column=3, row=2, padx=10, pady=10)
       
        boton_editar_detalle_venta.configure(command=lambda: self.editar_detalle_venta(detalles_treeview, entry_id_producto_detalle, 
                                                                                       entry_precio_producto_detalle, 
                                                                                       entry_cantidad_producto_detalle, lbl_total_venta))

        boton_borrar_detalle_venta = CTkButton(detalles_acciones, text="ELIMINAR\nPRODUCTO\nDE LA VENTA", 
                                        width=70, height=50, text_color='black', font=("OCR A Extended", 12),
                                        command=lambda: self.borrar_detalle_seleccionado(detalles_treeview,lbl_total_venta))
        boton_borrar_detalle_venta.grid(column=2, row=2, padx=10, pady=10)      


        separador1 = ttk.Separator(detalles_acciones, orient="vertical")
        separador1.grid(column=4, row=0,rowspan=3, sticky="ns",padx=5,pady=5)

        lbl_titulo01 = CTkLabel(detalles_acciones, text="TOTALES",
                                   font=("OCR A Extended",15),anchor="center")
        lbl_titulo01.grid(column=5, row=0,padx=5,pady=5)

        total_venta=0
        lbl_total_venta = CTkLabel(detalles_acciones, text="A pagar:\n${}".format(total_venta),
                           width=70, height=50, font=("OCR A Extended", 25),
                           anchor="center", bg_color="#F9F9FA", corner_radius=32)
        
        lbl_total_venta.grid(column=5, row=1,padx=5,pady=5)

        boton_grabar_venta = CTkButton(detalles_acciones, text="GRABAR\nVENTA",
                                       width=70, height=50, text_color='black',
                                       font=("OCR A Extended", 12),
                                       command=lambda: self.boton_grabar_venta(entry_busqueda_producto, treeview, lbl_pro_sel, lbl_pro_descripcion,
                       entry_precio_venta, entry_cantidad, lbl_subtotal_venta,
                       detalles_treeview, entry_precio_producto_detalle,
                       entry_nombre_producto_noinv, entry_cantidad_producto_detalle, lbl_total_venta,entry_valor_recibido,lbl_vuelto))
        boton_grabar_venta.grid(column=5, row=2,padx=5,pady=5)
        
        self.calcular_total_venta(detalles_treeview,lbl_total_venta)

        separador2 = ttk.Separator(detalles_acciones, orient="vertical")
        separador2.grid(column=6, row=0,rowspan=3, sticky="ns",padx=5,pady=5)

        lbl_valor_recibido = CTkLabel(detalles_acciones,text="Ingrese\nvalor\nrecibido:",width=70, height=50, text_color='black',
                                       font=("OCR A Extended", 15))
        lbl_valor_recibido.grid(column=7,row=0,padx=5,pady=5)

        entry_valor_recibido = CTkEntry(detalles_acciones,placeholder_text="\uf53d",width=70, height=50, text_color='black',
                                       font=("OCR A Extended", 15))
        entry_valor_recibido.grid(column=7,row=1,padx=5,pady=5)
        entry_valor_recibido.bind('<KeyRelease>', lambda event: self.calcular_vuelto(event,lbl_total_venta, entry_valor_recibido, lbl_vuelto))

        lbl_vuelto = CTkLabel(detalles_acciones,text="No hay\nventa registrada.",width=70, height=50, text_color='black',
                                       font=("OCR A Extended", 15))
        lbl_vuelto.grid(column=7,row=2,padx=5,pady=5)
        
    def calcular_vuelto(self, event, lbl_total_venta, entry_valor_recibido, lbl_vuelto):
        # Obtener el texto del total de la venta
        total_venta_text = lbl_total_venta.cget("text")
        
        # Validar si el texto del total de la venta está vacío
        if not total_venta_text:
            lbl_vuelto.configure(text="No hay venta registrada.")
            return
        
        # Obtener el valor numérico del total de la venta
        try:
            total_venta = float(total_venta_text.split("$")[-1])
        except ValueError:
            lbl_vuelto.configure(text="No hay\nventa registrada.")
            return

        # Obtener el valor recibido del Entry
        valor_recibido_text = entry_valor_recibido.get()
        if not valor_recibido_text:
            lbl_vuelto.configure(text="No hay\nventa registrada.")
            return

        try:
            # Obtener el valor recibido del Entry
            valor_recibido = float(valor_recibido_text)

            # Verificar si el valor recibido es mayor que el total de la venta
            if valor_recibido < total_venta:
                messagebox.showerror("Error", "El monto recibido debe ser mayor o igual al total de la venta.")
                return

            # Calcular el vuelto
            vuelto = valor_recibido - total_venta

            # Mostrar el vuelto en la etiqueta
            lbl_vuelto.configure(text="Vuelto: ${:.2f}".format(vuelto))

        except ValueError:
            messagebox.showerror("Error", "Por favor, ingrese un monto válido.")

    def calcular_total_venta(self,detalles_treeview,lbl_total_venta):
            # Obtener todas las filas del Treeview de detalles de venta
            filas = detalles_treeview.get_children()

            # Inicializar el total de la venta
            total_venta = 0

            # Iterar sobre todas las filas y sumar los subtotales
            for fila in filas:
                subtotal = float(detalles_treeview.item(fila, 'values')[-1])
                total_venta += subtotal
                total_venta = round(total_venta,2)

            # Mostrar el total de la venta en la etiqueta
            lbl_total_venta.configure(text="A pagar:\n${}".format(total_venta))

    def actualizar_precio_cantidad(self, event=None, entry_precio_venta=None, entry_cantidad=None, lbl_subtotal_venta=None):
        if entry_precio_venta and entry_cantidad and lbl_subtotal_venta:
            precio = entry_precio_venta.get()
            cantidad = entry_cantidad.get()

            try:
                if precio and cantidad:  
                    precio_float = float(precio)
                    cantidad_int = int(cantidad)
                    subtotal = precio_float * cantidad_int
                    subtotal = round(subtotal, 2)  # limite decimales
                    lbl_subtotal_venta.configure(text=f'subtotal\n$ {subtotal}')
                else:
                    lbl_subtotal_venta.configure(text='subtotal\n$ 0.0')
            except ValueError:
                lbl_subtotal_venta.configure(text='')
      
    def actualizar_precio_venta(self, event, entry_precio_venta, treeview, entry_cantidad, lbl_subtotal_venta, lbl_pro_sel, descripcion, lbl_pro_descripcion,entry_id_producto):
        # Obtener el ítem seleccionado en el Treeview
        selected_item = treeview.focus()
        if selected_item:
            # Obtener los valores de la fila seleccionada
            values = treeview.item(selected_item)['values']
            # El nombre del producto está en la primera posición (índice 0)
            id_producto = values[0]
            nombre_producto = values[1]
            # El precio está en la tercera posición (índice 2)
            precio = values[3]
            descripcion = values[2]

            entry_id_producto.delete(0,tk.END)
            entry_id_producto.insert(tk.END,id_producto)

            # Actualizar el entry de precio con el valor seleccionado
            entry_precio_venta.delete(0, tk.END)
            entry_precio_venta.insert(0, precio)

            # Actualizar la etiqueta con el nombre del producto seleccionado
            lbl_pro_sel.delete("1.0",tk.END)
            lbl_pro_sel.insert(tk.END,nombre_producto)

            # Limpiar el ScrolledText y luego insertar la nueva descripción
            lbl_pro_descripcion.delete("1.0", tk.END)  # Limpiar el ScrolledText
            lbl_pro_descripcion.insert(tk.END, descripcion)  # Insertar la nueva descripción

            entry_cantidad.delete(0, tk.END)
            entry_cantidad.insert(0, "1")
         
            self.actualizar_precio_cantidad(event, entry_precio_venta, entry_cantidad, lbl_subtotal_venta)

    def actualizar_filtrado(self, event, entry_busqueda_producto, treeview, lbl_pro_sel, lbl_pro_descripcion, entry_id_producto):
        # Obtener el texto de búsqueda
        texto_busqueda = entry_busqueda_producto.get()

        # Conectar a la base de datos y realizar la consulta
        conn = sqlite3.connect('farmacia.db')
        cursor = conn.cursor()

        # Limpiar el Treeview
        for item in treeview.get_children():
            treeview.delete(item)

        # Consultar la base de datos
        cursor.execute("SELECT id, nombre, descripcion, precio_venta, existencias, codigo_barras FROM productos WHERE nombre LIKE ? OR descripcion LIKE ? OR codigo_barras LIKE ?", ('%' + texto_busqueda + '%', '%' + texto_busqueda + '%', '%' + texto_busqueda + '%'))
        productos = cursor.fetchall()

        # Insertar los datos en el Treeview
        for producto in productos:
            treeview.insert('', 'end', values=producto)

        # Cerrar la conexión a la base de datos
        conn.close()

        # Verificar si hay algún producto seleccionado en el Treeview
        if not treeview.selection():
            # Si no hay ningún producto seleccionado, actualizar lbl_pro_sel y lbl_pro_descripcion
            lbl_pro_sel.delete("1.0", tk.END)  # Eliminar cualquier texto actual en lbl_pro_sel
            lbl_pro_sel.insert(tk.END, "NOMBRE")

            lbl_pro_descripcion.delete("1.0", tk.END)  # Limpiar el ScrolledText
            lbl_pro_descripcion.insert(tk.END, "DESCRIPCION")

            # También puedes limpiar el entry_id_producto si no hay selección
            entry_id_producto.delete(0, tk.END)
        else:
            # Si hay un producto seleccionado, obtener los valores y asignarlos a los widgets correspondientes
            selected_item = treeview.selection()
            values = treeview.item(selected_item)['values']

            # Asignar valores a los widgets
            entry_id_producto.delete(0, tk.END)
            entry_id_producto.insert(0, values[1])  # Asignar el ID al entry_id_producto

            lbl_pro_sel.delete("1.0", tk.END)
            lbl_pro_sel.insert(tk.END, values[2])  # Asignar el Nombre a lbl_pro_sel

            lbl_pro_descripcion.delete("1.0", tk.END)
            lbl_pro_descripcion.insert(tk.END, values[3])  # Asignar la Descripción a lbl_pro_descripcion

    def agregar_articulo(self, treeview, entry_cantidad, entry_precio_venta, entry_nombre_producto_noinv, detalles_treeview, lbl_total_venta):
        # Obtener el índice seleccionado en el Treeview de productos
        seleccion = treeview.selection()

        # Verificar si se ha seleccionado un producto desde el Treeview
        if seleccion:
            # Obtener los datos del producto seleccionado
            item = treeview.item(seleccion)
            id_producto_detalle = item['values'][0]
            nombre_producto = item['values'][1]
            precio_producto = float(item['values'][3])  # Convertir a float
        else:
            # Obtener los datos ingresados manualmente
            id_producto_detalle = self.generar_codigo_aleatorio()
            nombre_producto = entry_nombre_producto_noinv.get()
            precio_producto = entry_precio_venta.get()

        # Verificar si se proporcionó un nombre de producto y un precio válido
        if nombre_producto and precio_producto:
            try:
                # Verificar si el precio es un número positivo
                precio_producto = float(precio_producto)
                if precio_producto > 0:
                    # Obtener la cantidad especificada por el usuario
                    cantidad = entry_cantidad.get()

                    # Verificar si la cantidad es válida (es un número entero positivo)
                    if cantidad.isdigit() and int(cantidad) > 0:
                        # Verificar si el producto ya existe en el Treeview de detalles
                        productos_detalles = detalles_treeview.get_children()
                        producto_existente = False

                        for producto in productos_detalles:
                            valores = detalles_treeview.item(producto, 'values')
                            if valores[0] == id_producto_detalle:  # Comparar por ID en lugar de por nombre
                                producto_existente = True
                                break

                        if not producto_existente:
                            # Calcular el subtotal para el producto agregado
                            subtotal = int(cantidad) * float(precio_producto)
                            subtotal = round(subtotal, 2)

                            # Agregar una nueva fila al Treeview de detalles de venta con la información del producto y la cantidad
                            detalles_treeview.insert("", "end", values=(id_producto_detalle, nombre_producto, precio_producto, cantidad, subtotal))

                            # Limpiar los campos de entrada
                            entry_cantidad.delete(0, 'end')
                            entry_nombre_producto_noinv.delete(0, 'end')
                            entry_precio_venta.delete(0, 'end')

                            # Calcular y mostrar el total de la venta
                            self.calcular_total_venta(detalles_treeview, lbl_total_venta)
                            treeview.focus_set()
                        else:
                            # Mostrar un mensaje de error si el producto ya existe en detalles_treeview
                            messagebox.showerror("Error", "El producto ya existe en la lista de detalles, puedes editar ahí el precio y la cantidad a vender.")
                    else:
                        # Mostrar un mensaje de error si la cantidad no es válida
                        messagebox.showerror("Error", "La cantidad ingresada no es válida o no es un número entero positivo.")
                else:
                    # Mostrar un mensaje de error si el precio no es positivo
                    messagebox.showerror("Error", "El precio ingresado debe ser un número positivo.")
            except ValueError:
                # Mostrar un mensaje de error si el precio no es un número válido
                messagebox.showerror("Error", "El precio ingresado no es válido.")
        else:
            # Mostrar un mensaje de error si no se proporcionó nombre de producto o precio
            messagebox.showerror("Error", "Por favor, ingrese nombre de producto y precio antes de agregarlo a la venta.")

    def actualizar_campos_entrada(self, detalles_treeview,id_det,entry_id_producto_detalle, entry_precio_producto_detalle, entry_cantidad_producto_detalle):
        selected_item = detalles_treeview.focus()

        if selected_item:
            values = detalles_treeview.item(selected_item)['values']
            id_det = values[0]
            nombre_producto = values[1]
            cantidad = values[2]
            precio = values[3]


            entry_id_producto_detalle.delete(0, 'end')
            entry_cantidad_producto_detalle.delete(0, 'end')
            entry_precio_producto_detalle.delete(0, 'end')

            entry_id_producto_detalle.insert(0, id_det)
            entry_cantidad_producto_detalle.insert(0, cantidad)
            entry_precio_producto_detalle.insert(0, precio)

    def borrar_detalle_seleccionado(self, detalles_treeview,lbl_total_venta):
        selected_item = detalles_treeview.selection()

        if selected_item:
            detalles_treeview.delete(selected_item)
            
            self.calcular_total_venta(detalles_treeview,lbl_total_venta)
    
    def editar_detalle_venta(self, detalles_treeview, entry_id_producto_detalle, entry_precio_producto_detalle, entry_cantidad_producto_detalle, lbl_total_venta):
        # Obtener la fila seleccionada en el Treeview de detalles
        selected_item = detalles_treeview.selection()

        if selected_item:
            try:
                # Obtener los nuevos valores desde los campos de entrada
                id_producto_detalle = entry_id_producto_detalle.get()
                producto_original = detalles_treeview.item(selected_item, 'values')[1]
                cantidad = entry_cantidad_producto_detalle.get()
                precio = float(entry_precio_producto_detalle.get())

                # Validar que los nuevos valores sean válidos
                if cantidad and isinstance(float(cantidad), float) and float(cantidad) > 0 and precio > 0:

                    # Actualizar los valores en el Treeview de detalles
                    subtotal = round(float(cantidad) * precio, 2)
                    detalles_treeview.item(selected_item, values=(id_producto_detalle, producto_original, cantidad, precio, subtotal))

                    self.calcular_total_venta(detalles_treeview,lbl_total_venta)

                    # Limpiar los campos de entrada
                    entry_id_producto_detalle.delete(0, 'end')
                    entry_cantidad_producto_detalle.delete(0, 'end')
                    entry_precio_producto_detalle.delete(0, 'end')
                else:
                    # Mostrar un mensaje de error si los nuevos valores no son válidos
                    messagebox.showerror("Error", "Por favor, ingrese valores válidos para la cantidad y el precio.")
            except ValueError:
                # Mostrar un mensaje de error si los nuevos valores no son números válidos
                messagebox.showerror("Error", "Por favor, ingrese valores numéricos válidos para la cantidad y el precio.")
        else:
            # Mostrar un mensaje de error si no se seleccionó ninguna fila para editar
            messagebox.showerror("Error", "Por favor, seleccione una fila para editar en el Treeview de detalles.")

    def boton_grabar_venta(self, entry_busqueda_producto, treeview, lbl_pro_sel, lbl_pro_descripcion,
                        entry_precio_venta, entry_cantidad, lbl_subtotal_venta,
                        detalles_treeview, entry_precio_producto_detalle,
                        entry_nombre_producto_noinv, entry_cantidad_producto_detalle, 
                        lbl_total_venta, entry_valor_recibido, lbl_vuelto):

        # Obtener la fecha y hora actual
        fecha_venta = datetime.now().strftime("%Y-%m-%d")
        hora_venta = datetime.now().strftime("%H:%M:%S")
        metodo_pago = "Contado"
        id_cliente = self.generar_codigo_aleatorio()
        cliente = "consumidor final"  # Aquí deberías obtener el cliente adecuado, por ejemplo, seleccionándolo de una lista

        # Conectar a la base de datos
        conn = sqlite3.connect('farmacia.db')
        cursor = conn.cursor()

        try:
            # Verificar si hay una jornada en curso para la fecha actual
            cursor.execute("SELECT * FROM operaciones_caja WHERE fecha_inicio = ? AND estado = ?", (fecha_venta, "Jornada en curso"))
            jornada_en_curso = cursor.fetchone()

            if not jornada_en_curso:
                messagebox.showwarning("Aviso", "No hay una jornada en curso para el día de hoy. No se puede grabar la venta, INICIE OPERACIONES.")
                self.formulario_inicio_operaciones()
                return

            # Si hay una jornada en curso, procede a grabar la venta
            # Crear la tabla de ventas si no existe
            cursor.execute('''CREATE TABLE IF NOT EXISTS ventas (
                    id TEXT PRIMARY KEY,
                    id_producto TEXT,
                    producto TEXT,
                    id_cliente TEXT,
                    cliente TEXT,
                    cantidad INTEGER,
                    precio_unitario REAL,
                    total REAL,
                    metodo_pago TEXT,
                    fecha DATE,
                    hora TIME,
                    categoria_producto TEXT,
                    FOREIGN KEY (id_producto) REFERENCES productos (id)
                )''')

            # Obtener todas las filas del Treeview de detalles de venta
            filas = detalles_treeview.get_children()

            # Iterar sobre todas las filas y grabar cada detalle de venta en la base de datos
            for fila in filas:
                valores = detalles_treeview.item(fila, 'values')
                id_detalle_venta = self.generar_codigo_aleatorio()  # Generar un nuevo ID de venta para cada detalle
                id_producto = valores[0]  # Obtener el ID del producto
                nombre_producto = valores[1]
                cantidad = float(valores[3])  # Obtener la cantidad
                precio_unitario = float(valores[2])
                subtotal = float(valores[4])  # Obtener el subtotal
                
                # Verificar si el producto está en la base de datos
                cursor.execute("SELECT existencias FROM productos WHERE id = ?", (id_producto,))
                existencias_result = cursor.fetchone()
                if existencias_result is None:
                    existencias_actuales = 0  # Establecer existencias actuales como 0 si el producto no está en la base de datos
                    
                    # Establecer la categoría predeterminada como "Producto No inventariado"
                    categoria_producto = "Producto No inventariado"
                else:
                    existencias_actuales = existencias_result[0]
                    
                    # Consultar la categoría del producto
                    cursor.execute("SELECT categoria FROM productos WHERE id = ?", (id_producto,))
                    categoria_result = cursor.fetchone()
                    categoria_producto = categoria_result[0] if categoria_result else None
                
                # Calcular las nuevas existencias después de la venta
                nuevas_existencias = existencias_actuales - cantidad

                # Actualizar las existencias del producto en la base de datos
                cursor.execute("UPDATE productos SET existencias = ? WHERE id = ?", (nuevas_existencias, id_producto))

                # Insertar el detalle de venta en la tabla de ventas
                cursor.execute('''
                    INSERT INTO ventas (id, id_producto, producto, id_cliente, cliente, cantidad, precio_unitario, total, metodo_pago, fecha, hora, categoria_producto)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (id_detalle_venta, id_producto, nombre_producto, id_cliente, cliente, cantidad, precio_unitario, subtotal, metodo_pago, fecha_venta, hora_venta, categoria_producto))

            # Confirmar la transacción
            conn.commit()
            
            # Limpiar el Treeview de detalles de venta después de grabar la venta
            detalles_treeview.delete(*filas)
            
            # Consultar nuevamente la base de datos para obtener los productos actualizados
            cursor.execute("SELECT id, nombre, descripcion, precio_venta, existencias, codigo_barras FROM productos")
            productos_actualizados = cursor.fetchall()

            # Limpiar el Treeview
            treeview.delete(*treeview.get_children())

            # Insertar los datos actualizados en el Treeview
            for producto in productos_actualizados:
                treeview.insert('', 'end', values=producto)

            # Cerrar la conexión a la base de datos
            conn.close()
            
            # Restablecer valores
            self.restablecer_valores(entry_busqueda_producto, treeview, lbl_pro_sel, lbl_pro_descripcion,
                                entry_precio_venta, entry_cantidad, lbl_subtotal_venta,
                                detalles_treeview, entry_precio_producto_detalle,
                                entry_nombre_producto_noinv, entry_cantidad_producto_detalle, 
                                lbl_total_venta, entry_valor_recibido, lbl_vuelto)

            # Mostrar un mensaje de éxito
            messagebox.showinfo("Venta grabada", "La venta se ha grabado exitosamente en la base de datos.")
            
        except sqlite3.Error as e:
            # Manejar cualquier error en la operación de la base de datos
            messagebox.showerror("Error", f"No se pudo grabar la venta en la base de datos.\nError: {e}")

            # Cerrar la conexión en caso de error
            conn.close()

    def generar_codigo_aleatorio(self):
        # Definir la longitud del código
        longitud = 8

        # Crear una lista que contenga dígitos y letras
        caracteres = string.digits + string.ascii_letters

        # Generar el código aleatorio utilizando random.choice
        codigo_aleatorio = ''.join(random.choice(caracteres) for _ in range(longitud))

        return codigo_aleatorio
    
    def deselect_item(self,event,lbl_pro_sel,lbl_pro_descripcion,treeview,entry_precio_venta,entry_cantidad,entry_busqueda_producto,lbl_subtotal_venta):
        treeview.selection_remove(treeview.selection())
        lbl_pro_sel.delete(1.0, tk.END)
        lbl_pro_sel.insert(tk.END,"NOMBRE")
        lbl_pro_descripcion.delete(1.0, tk.END)
        lbl_pro_descripcion.insert(tk.END, "DESCRIPCION")
        entry_precio_venta.delete(0, tk.END)
        entry_precio_venta._activate_placeholder()  
        entry_cantidad.delete(0, tk.END)
        entry_cantidad._activate_placeholder()  
        entry_busqueda_producto.delete(0, tk.END)
        entry_busqueda_producto._activate_placeholder()
        lbl_subtotal_venta.configure(text="subtotal\n$")

    def restablecer_valores(self, entry_busqueda_producto, treeview, lbl_pro_sel, lbl_pro_descripcion,
                        entry_precio_venta, entry_cantidad, lbl_subtotal_venta,
                        detalles_treeview, entry_precio_producto_detalle,entry_nombre_producto_noinv, 
                        entry_cantidad_producto_detalle, lbl_total_venta,entry_valor_recibido,lbl_vuelto):
        # Restablecer valores de Entry

        entry_busqueda_producto.delete(0, tk.END)
        entry_precio_venta.delete(0, tk.END)
        entry_cantidad.delete(0, tk.END)
        entry_cantidad._activate_placeholder()
        
        lbl_vuelto.configure(text="No hay\nventa registrada.")
        entry_valor_recibido.delete(0, tk.END)
        entry_valor_recibido._activate_placeholder()
        
        entry_precio_producto_detalle.delete(0, tk.END)
        entry_cantidad_producto_detalle.delete(0, tk.END)
        entry_precio_producto_detalle._activate_placeholder()
        entry_cantidad_producto_detalle._activate_placeholder()
        # Restablecer valores 
        lbl_pro_sel.delete(1.0, tk.END)
        lbl_pro_sel.insert(tk.END,"NOMBRE")
        lbl_pro_descripcion.delete(1.0, tk.END)
        lbl_pro_descripcion.insert(tk.END, "DESCRIPCION")
        lbl_subtotal_venta.configure(text="subtotal\n$")
        
        treeview.selection_remove(treeview.selection())    

        # Restablecer valores de Treeview de detalles
        detalles_treeview.delete(*detalles_treeview.get_children())
        entry_busqueda_producto._activate_placeholder()
        entry_precio_venta._activate_placeholder()
        # Restablecer valores de lbl_total_venta
        lbl_total_venta.configure(text="A pagar:\n$")  # Puedes establecer otro valor por defecto si es necesario
        entry_nombre_producto_noinv._activate_placeholder()

    def on_treeview_select(self, event, detalles_treeview, id_det, entry_id_producto_detalle, entry_precio_producto_detalle, entry_cantidad_producto_detalle, lbl_pro_sel, lbl_pro_descripcion, treeview, entry_precio_venta, entry_cantidad, entry_busqueda_producto, lbl_subtotal_venta):
        # Update entry fields with selected item data
        self.actualizar_campos_entrada(detalles_treeview, id_det, entry_id_producto_detalle, entry_precio_producto_detalle, entry_cantidad_producto_detalle)
        
        # Deselect item and update related fields
        self.deselect_item(event, lbl_pro_sel, lbl_pro_descripcion, treeview, entry_precio_venta, entry_cantidad, entry_busqueda_producto, lbl_subtotal_venta)
    
    def move_selection_detalles(self,event,detalles_treeview):
        if event.keysym == 'Up':
            detalles_treeview.focus_set()
            detalles_treeview.selection_add(detalles_treeview.prev(detalles_treeview.selection()))
        elif event.keysym == 'Down':
            detalles_treeview.focus_set()
            detalles_treeview.selection_add(detalles_treeview.next(detalles_treeview.selection()))
     
    def move_selection_filtrado(self, event, treeview, entry_cantidad):
        if event.keysym == 'Up':
            treeview.focus_set()
            treeview.selection_add(treeview.prev(treeview.selection()))
        elif event.keysym == 'Down':
            treeview.focus_set()
            treeview.selection_add(treeview.next(treeview.selection()))
        elif event.keysym == 'Return':  # Verificar si se presionó Enter
            entry_cantidad.focus()  # Hacer que el cursor caiga en el entry de cantidad
    
    def formulario_inicio_operaciones(self):
        
        ESTILO_CTKBOTONES = {
            'width': 50,
            'height': 40,
            'text_color': 'black',
            'font': ("OCR A Extended", 13)
        }
       
        ESTILO_ENTRYS_LABEL = {
            'text_color': 'black',
            'font': ("OCR A Extended", 14),
               
        }
        
        ESTILO_TITULO = {
            
            'font': ("OCR A Extended", 16,"bold"),  

        }
        
        conexion = sqlite3.connect('farmacia.db')
        cursor = conexion.cursor()

        # Verificar si ya existe un registro para la fecha actual
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("SELECT COUNT(*) FROM operaciones_caja WHERE fecha_inicio = ?", (fecha_actual,))
        existe_registro = cursor.fetchone()[0]

        if existe_registro:
            messagebox.showinfo("Ya se inicio la caja", "La jornada de trabajo esta en curso.")
            return
    
        
        # Crear una nueva ventana
        self.inicio_operaciones = tk.Toplevel(self.cuerpo_principal,background=COLOR_CUERPO_PRINCIPAL)
        self.inicio_operaciones.title("Inicio de Operaciones")
        self.inicio_operaciones.iconbitmap("./imagenes/logo.ico")
        #self.inicio_operaciones.geometry("400x300")
        self.inicio_operaciones.resizable(False, False)

        # Crear el formulario dentro de la ventana
        frame_formulario_inicio_operaciones = tk.Frame(self.inicio_operaciones,background=COLOR_CUERPO_PRINCIPAL)
        frame_formulario_inicio_operaciones.pack(padx=10, pady=10)

        # Etiquetas y campos de entrada del formulario
        lbl_titulo_inicio = CTkLabel(frame_formulario_inicio_operaciones, text="Inicio de Operaciones",**ESTILO_TITULO)
        lbl_titulo_inicio.grid(column=0, row=0, columnspan=2, pady=5)

        entry_id_inicio_operaciones=tk.Entry(frame_formulario_inicio_operaciones)
        entry_id_inicio_operaciones.grid_forget
        codigo_aleatorio = self.generar_codigo_aleatorio() 
        entry_id_inicio_operaciones.insert(0,codigo_aleatorio)

        entry_estado_inicial=tk.Entry(frame_formulario_inicio_operaciones)
        entry_estado_inicial.grid_forget
        entry_estado_inicial.insert(0,"Jornada en curso")

        lbl_fecha_inicio = CTkLabel(frame_formulario_inicio_operaciones, text="Fecha de Inicio:",**ESTILO_ENTRYS_LABEL)
        lbl_fecha_inicio.grid(column=0, row=1, padx=5, pady=5, sticky='e')

        entry_fecha_inicio = CTkEntry(frame_formulario_inicio_operaciones,**ESTILO_ENTRYS_LABEL)
        entry_fecha_inicio.grid(column=1, row=1, padx=5, pady=5, sticky='w')
        entry_fecha_inicio.insert(0, datetime.now().strftime('%Y-%m-%d'))
        entry_fecha_inicio.configure(state=tk.DISABLED)

        lbl_hora_inicio = CTkLabel(frame_formulario_inicio_operaciones, text="Hora de Inicio:",**ESTILO_ENTRYS_LABEL)
        lbl_hora_inicio.grid(column=0, row=2, padx=5, pady=5, sticky='e')

        entry_hora_inicio = CTkEntry(frame_formulario_inicio_operaciones,**ESTILO_ENTRYS_LABEL)
        entry_hora_inicio.grid(column=1, row=2, padx=5, pady=5, sticky='w')
        entry_hora_inicio.insert(0, datetime.now().strftime('%H:%M:%S'))
        entry_hora_inicio.configure(state=tk.DISABLED)

        lbl_usuario_logueado = CTkLabel(frame_formulario_inicio_operaciones, text="Usuario:",**ESTILO_ENTRYS_LABEL)
        lbl_usuario_logueado.grid(column=0, row=3, padx=5, pady=5, sticky='e')

        entry_usuario_logueado = CTkEntry(frame_formulario_inicio_operaciones,**ESTILO_ENTRYS_LABEL)
        entry_usuario_logueado.grid(column=1, row=3, padx=5, pady=5, sticky='w')
        entry_usuario_logueado.insert(0,"Administrador")
        entry_usuario_logueado.configure(state=tk.DISABLED)

        lbl_valor_inicio = CTkLabel(frame_formulario_inicio_operaciones, text="Ingrese Valor Inicial:",**ESTILO_ENTRYS_LABEL)
        lbl_valor_inicio.grid(column=0, row=4, padx=5, pady=5, sticky='e')
    
        entry_valor_inicial = CTkEntry(frame_formulario_inicio_operaciones,placeholder_text="\uf53d",**ESTILO_ENTRYS_LABEL)
        entry_valor_inicial.grid(column=1, row=4, padx=5, pady=5, sticky='w')
    
        btn_confirmar = CTkButton(frame_formulario_inicio_operaciones,text="Confirmar Inicio",command=lambda :self.guardar_operacion(entry_id_inicio_operaciones, 
                                                                                                                                     entry_fecha_inicio, entry_hora_inicio, 
                                                                                                                                     entry_usuario_logueado, entry_valor_inicial,
                                                                                                                                     entry_estado_inicial),**ESTILO_CTKBOTONES)
        btn_confirmar.grid(column=1, row=5, padx=5, pady=5, sticky='w')
    
    def guardar_operacion(self, entry_id_inicio_operaciones, entry_fecha_inicio, entry_hora_inicio, entry_usuario_logueado, entry_valor_inicial,entry_estado_inicial):
        try:
            # Conexión a la base de datos
            conexion = sqlite3.connect('farmacia.db')
            cursor = conexion.cursor()

            # Obtener los valores ingresados por el usuario
            id_inicio_operaciones = entry_id_inicio_operaciones.get()
            fecha_inicio = entry_fecha_inicio.get()
            hora_inicio = entry_hora_inicio.get()
            estado_inicial = entry_estado_inicial.get()
            usuario = entry_usuario_logueado.get()
            valor_inicial = entry_valor_inicial.get()

            # Insertar los datos en la tabla operaciones_caja
            cursor.execute('''INSERT INTO operaciones_caja 
                            (id, fecha_inicio, hora_inicio, usuario, valor_inicial,estado) 
                            VALUES (?, ?, ?, ?, ?,?)''', 
                            (id_inicio_operaciones, fecha_inicio, hora_inicio, usuario, float(valor_inicial),estado_inicial))
            
            conexion.commit()  # Confirmar la transacción

            # Mostrar un mensaje de confirmación
            messagebox.showinfo("Operación Exitosa", "Los datos se han guardado correctamente.")

        except Exception as e:
            # En caso de error, deshacer cualquier cambio en la base de datos
            conexion.rollback()
            # Mostrar un mensaje de error
            messagebox.showerror("Error", f"No se pudo guardar la operación. Error: {str(e)}")

        finally:
            # Cerrar la conexión a la base de datos
            if conexion:
                conexion.close()

    def formulario_cierre_caja(self):
        ESTILO_CTKBOTONES = {
            'width': 50,
            'height': 50,  # Ajustado para botones más grandes
            'text_color': 'black',
            'font': ("OCR A Extended", 13)
        }
    
        ESTILO_ENTRYS_LABEL = {
            'text_color': 'black',
            'font': ("OCR A Extended", 14)
        }
        
        ESTILO_TITULO = {
            'font': ("OCR A Extended", 16, "bold")
        }

        # Crear una nueva ventana
        self.cierre_caja = tk.Toplevel(self.cuerpo_principal, background=COLOR_CUERPO_PRINCIPAL)
        self.cierre_caja.title("Cierre de caja")
        self.cierre_caja.iconbitmap("./imagenes/logo.ico")
        #self.cierre_caja.geometry("500x600")
        self.cierre_caja.resizable(False, False)

        # Crear el formulario dentro de la ventana
        frame_formulario_cierre_caja = tk.Frame(self.cierre_caja, background=COLOR_CUERPO_PRINCIPAL)
        frame_formulario_cierre_caja.pack(padx=10, pady=10)

        # Etiquetas y campos de entrada del formulario
        lbl_titulo_cierre = CTkLabel(frame_formulario_cierre_caja, text="Cierre de caja", **ESTILO_TITULO)
        lbl_titulo_cierre.grid(column=0, row=0, columnspan=2, pady=5)

        lbl_fecha_cierre = CTkLabel(frame_formulario_cierre_caja, text="Fecha de Inicio:", **ESTILO_ENTRYS_LABEL)
        lbl_fecha_cierre.grid(column=0, row=1, padx=5, pady=5, sticky='e')

        entry_fecha_cierre = CTkEntry(frame_formulario_cierre_caja, **ESTILO_ENTRYS_LABEL)
        entry_fecha_cierre.grid(column=1, row=1, padx=5, pady=5, sticky='w')
        entry_fecha_cierre.insert(0, datetime.now().strftime('%Y-%m-%d'))
        entry_fecha_cierre.configure(state=tk.DISABLED)

        lbl_hora_cierre = CTkLabel(frame_formulario_cierre_caja, text="Hora de Inicio:", **ESTILO_ENTRYS_LABEL)
        lbl_hora_cierre.grid(column=0, row=2, padx=5, pady=5, sticky='e')

        entry_hora_cierre = CTkEntry(frame_formulario_cierre_caja, **ESTILO_ENTRYS_LABEL)
        entry_hora_cierre.grid(column=1, row=2, padx=5, pady=5, sticky='w')
        entry_hora_cierre.insert(0, datetime.now().strftime('%H:%M:%S'))
        entry_hora_cierre.configure(state=tk.DISABLED)

        lbl_usuario_logueado = CTkLabel(frame_formulario_cierre_caja, text="Usuario:", **ESTILO_ENTRYS_LABEL)
        lbl_usuario_logueado.grid(column=0, row=3, padx=5, pady=5, sticky='e')

        entry_usuario_logueado = CTkEntry(frame_formulario_cierre_caja, **ESTILO_ENTRYS_LABEL)
        entry_usuario_logueado.grid(column=1, row=3, padx=5, pady=5, sticky='w')
        entry_usuario_logueado.insert(0,"Administrador")
        entry_usuario_logueado.configure(state=tk.DISABLED)

       # Crear un Frame para contener el Treeview y el Scrollbar
        tree_frame_movimientos = tk.Frame(frame_formulario_cierre_caja)
        tree_frame_movimientos.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")  # Ajuste de columnspan a 2

        # Configurar el estilo
        style = ttk.Style()
        style.configure("mystyle.Treeview", highlightthickness=0, bd=0, font=("OCR A Extended", 10))
        style.configure("mystyle.Treeview.Heading", font=("OCR A Extended", 9, "bold"))
        style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])

        # Crear un Treeview en el frame con el estilo configurado
        movimientos_treeview = ttk.Treeview(tree_frame_movimientos, style="mystyle.Treeview",height=4)
        movimientos_treeview.grid(row=0, column=0, sticky="nsew")

        # Configurar las columnas del Treeview
        movimientos_treeview["columns"] = ("ID", "Tipo", "Fecha", "Hora", "Monto", "Usuario", "Detalles")
        movimientos_treeview.heading("#0", text="ID", anchor="center")
        movimientos_treeview.heading("ID", text="ID")
        movimientos_treeview.heading("Tipo", text="Tipo")
        movimientos_treeview.heading("Fecha", text="Fecha")
        movimientos_treeview.heading("Hora", text="Hora")
        movimientos_treeview.heading("Monto", text="Monto")
        movimientos_treeview.heading("Usuario", text="Usuario")
        movimientos_treeview.heading("Detalles", text="Detalles")

        # Ajustar el ancho de las columnas
        movimientos_treeview.column("#0", width=0, stretch=tk.NO)
        movimientos_treeview.column("ID", width=0, stretch=tk.NO)
        movimientos_treeview.column("Tipo", width=100, anchor="center")
        movimientos_treeview.column("Fecha", width=100, anchor="center")
        movimientos_treeview.column("Hora", width=100, anchor="center")
        movimientos_treeview.column("Monto", width=100, anchor="center")
        movimientos_treeview.column("Usuario", width=100, anchor="center")
        movimientos_treeview.column("Detalles", width=150, anchor="center")
        self.ver_movimientos_caja(movimientos_treeview)
        
        valor_inicio_caja = CTkLabel(frame_formulario_cierre_caja, text="Valor Inicial: ",**ESTILO_ENTRYS_LABEL)
        valor_inicio_caja.grid(column=0,row=6,padx=5, pady=5, sticky='e')

        lbl_total_entradas = CTkLabel(frame_formulario_cierre_caja,text="valor Total Entrada:",**ESTILO_ENTRYS_LABEL)
        lbl_total_entradas.grid(column=0,row=7, padx=5, pady=5, sticky='e')

        lbl_total_ventas_diarias = CTkLabel(frame_formulario_cierre_caja,text="Total Ventas Diarias:",**ESTILO_ENTRYS_LABEL)
        lbl_total_ventas_diarias.grid(column=0,row=8, padx=5, pady=5, sticky='e')

        lbl_total_salidas = CTkLabel(frame_formulario_cierre_caja,text="Valor  Total Salidas: ", **ESTILO_ENTRYS_LABEL)
        lbl_total_salidas.grid(column=0,row=9, padx=5, pady=5, sticky='e')

        valor_presente_caja = CTkLabel(frame_formulario_cierre_caja,**ESTILO_ENTRYS_LABEL)
        valor_presente_caja.grid(column=0, row=10, padx=5, pady=5, sticky='w')

        entry_valor_presente = tk.Entry(frame_formulario_cierre_caja,text="")
        entry_valor_presente.grid_forget

        lbl_valor_cierre = CTkLabel(frame_formulario_cierre_caja, text="Ingrese Valor de cierre:", **ESTILO_ENTRYS_LABEL)
        lbl_valor_cierre.grid(column=0, row=11, padx=5, pady=5, sticky='e')

        entry_valor_cierre = CTkEntry(frame_formulario_cierre_caja, placeholder_text="\uf53d", **ESTILO_ENTRYS_LABEL)
        entry_valor_cierre.grid(column=1, row=11, padx=5, pady=5, sticky='w')
        
        entry_valor_cierre.bind("<Return>", lambda event: self.calculo_cierre(entry_valor_presente, entry_valor_cierre, informacion_final))
        
        self.actualizar_datos_cierre_jornada(lbl_total_salidas, lbl_total_entradas, lbl_total_ventas_diarias,valor_inicio_caja,
                                             valor_presente_caja,entry_valor_presente)

        btn_confirmar = CTkButton(frame_formulario_cierre_caja, text="Confirmar\ncierre de jornada", **ESTILO_CTKBOTONES)
        btn_confirmar.grid(column=1, row=12, padx=5, pady=5)  

        informacion_final = CTkLabel(frame_formulario_cierre_caja,text="",**ESTILO_ENTRYS_LABEL)
        informacion_final.grid(column=0, row=12, padx=5, pady=5)
        entry_valor_cierre.bind("<Return>", lambda event: self.calculo_cierre(entry_valor_presente, entry_valor_cierre, informacion_final))

        
    
    
    def actualizar_datos_cierre_jornada(self, lbl_total_salidas, lbl_total_entradas, lbl_total_ventas_diarias, valor_inicio_caja, 
                                        valor_presente_caja,entry_valor_presente):
        try:
            conexion = sqlite3.connect('farmacia.db')
            cursor = conexion.cursor()

            # Obtener la fecha actual en formato YYYY-MM-DD
            fecha_actual = datetime.now().strftime('%Y-%m-%d')

            # Obtener el valor inicial de la caja para la fecha actual
            cursor.execute("SELECT valor_inicial FROM operaciones_caja WHERE fecha_inicio = ?", (fecha_actual,))
            valor_inicial = cursor.fetchone()[0] or 0.0

            # Calcular el total de salidas para la fecha actual
            cursor.execute("SELECT SUM(monto) FROM movimientos_caja WHERE fecha = ? AND tipo = ?", (fecha_actual, "Salida de dinero"))
            total_salidas = cursor.fetchone()[0] or 0.0

            # Calcular el total de entradas para la fecha actual
            cursor.execute("SELECT SUM(monto) FROM movimientos_caja WHERE fecha = ? AND tipo = ?", (fecha_actual, "Entrada de dinero"))
            total_entradas = cursor.fetchone()[0] or 0.0

            # Calcular el total de ventas diarias
            cursor.execute("SELECT SUM(total) FROM ventas WHERE fecha = ?", (fecha_actual,))
            total_ventas_diarias = cursor.fetchone()[0] or 0.0

            # Calcular el valor final de la caja
            valor_final_caja = valor_inicial + total_entradas + total_ventas_diarias - total_salidas
            valor_final_redondeado = round(valor_final_caja, 2)

            # Actualizar los labels con los valores calculados
            valor_inicio_caja.configure(text=f"Valor Inicio caja: $ {valor_inicial}")
            lbl_total_salidas.configure(text=f"Valor Total Salidas: $ {total_salidas}")
            lbl_total_entradas.configure(text=f"Valor Total Entradas: $ {total_entradas}")
            lbl_total_ventas_diarias.configure(text=f"Total Ventas Diarias: $ {total_ventas_diarias}")
            valor_presente_caja.configure(text=f"El valor que debe tener en caja es: $ {valor_final_redondeado}")
            entry_valor_presente.insert(0, valor_final_redondeado)  # Insertar el nuevo valor
            
            conexion.commit()

        except sqlite3.Error as e:
            if conexion:
                conexion.rollback()
            messagebox.showerror("Error", f"No se pudo obtener los datos de cierre de jornada. Error: {str(e)}")

        finally:
            if conexion:
                conexion.close()

    def calculo_cierre(self, entry_valor_presente, entry_valor_cierre, informacion_final):
        # Obtener los valores como cadenas de texto desde los widgets de entrada
        valor_final_redondeado_str = entry_valor_presente.get()
        valor_cierre_str = entry_valor_cierre.get()

        try:
            # Convertir las cadenas de texto a números decimales
            valor_final_redondeado = float(valor_final_redondeado_str)
            valor_cierre = float(valor_cierre_str)

            # Calcular la diferencia entre los valores
            diferencia = valor_cierre - valor_final_redondeado

            # Determinar el mensaje según la diferencia
            if diferencia < 0:
                mensaje = f"Hay un faltante de caja de ${abs(diferencia)}."
            elif diferencia > 0:
                mensaje = f"Hay un superávit de caja de ${diferencia}."
            else:
                mensaje = "El cierre de caja coincide con el valor de cierre."

            # Mostrar el mensaje en el widget informacion_final
            informacion_final.configure(text=mensaje)

        except ValueError:
            # Manejar el caso en el que la conversión a decimal falle
            messagebox.showerror("Error", "Por favor, ingrese números válidos en los campos de entrada.")


    def formulario_movimientos_caja(self):
        ESTILO_CTKBOTONES = {
            'width': 50,
            'height': 50,  # Ajustado para botones más grandes
            'text_color': 'black',
            'font': ("OCR A Extended", 13)
        }
    
        ESTILO_ENTRYS_LABEL = {
            'text_color': 'black',
            'font': ("OCR A Extended", 14)
        }
        
        ESTILO_TITULO = {
            'font': ("OCR A Extended", 16, "bold")
        }

        # Crear una nueva ventana
        self.movimientos_caja = tk.Toplevel(self.cuerpo_principal, background=COLOR_CUERPO_PRINCIPAL)
        self.movimientos_caja.title("Movimientos en la caja")
        self.movimientos_caja.iconbitmap("./imagenes/logo.ico")
        self.movimientos_caja.resizable(False, False)

        # Crear el formulario dentro de la ventana
        frame_formulario_movimientos = tk.Frame(self.movimientos_caja, background=COLOR_CUERPO_PRINCIPAL)
        frame_formulario_movimientos.pack(padx=10, pady=10)

        # Etiquetas y campos de entrada del formulario
        lbl_titulo = CTkLabel(frame_formulario_movimientos, text="Movimientos caja", **ESTILO_TITULO)
        lbl_titulo.grid(column=0, row=0, columnspan=2, pady=5)

        lbl_fecha = CTkLabel(frame_formulario_movimientos, text="Fecha de Inicio:", **ESTILO_ENTRYS_LABEL)
        lbl_fecha.grid(column=0, row=1, padx=5, pady=5, sticky='e')

        entry_fecha = CTkEntry(frame_formulario_movimientos, **ESTILO_ENTRYS_LABEL)
        entry_fecha.grid(column=1, row=1, padx=5, pady=5, sticky='w')
        entry_fecha.insert(0, datetime.now().strftime('%Y-%m-%d'))
        entry_fecha.configure(state=tk.DISABLED)

        lbl_hora = CTkLabel(frame_formulario_movimientos, text="Hora de Inicio:", **ESTILO_ENTRYS_LABEL)
        lbl_hora.grid(column=0, row=2, padx=5, pady=5, sticky='e')

        entry_hora = CTkEntry(frame_formulario_movimientos, **ESTILO_ENTRYS_LABEL)
        entry_hora.grid(column=1, row=2, padx=5, pady=5, sticky='w')
        entry_hora.insert(0, datetime.now().strftime('%H:%M:%S'))
        entry_hora.configure(state=tk.DISABLED)

        lbl_usuario_logueado = CTkLabel(frame_formulario_movimientos, text="Usuario:", **ESTILO_ENTRYS_LABEL)
        lbl_usuario_logueado.grid(column=0, row=3, padx=5, pady=5, sticky='e')

        entry_usuario_logueado = CTkEntry(frame_formulario_movimientos, **ESTILO_ENTRYS_LABEL)
        entry_usuario_logueado.grid(column=1, row=3, padx=5, pady=5, sticky='w')
        entry_usuario_logueado.insert(0,"Administrador")
        entry_usuario_logueado.configure(state=tk.DISABLED)

        entry_tipo = CTkOptionMenu(frame_formulario_movimientos,values=["Salida de dinero", "Entrada de dinero"],**ESTILO_ENTRYS_LABEL)
        entry_tipo.grid(column=0, row=6, padx=5, pady=5, sticky='e')

        entry_monto = CTkEntry(frame_formulario_movimientos, placeholder_text="\uf53d", **ESTILO_ENTRYS_LABEL)
        entry_monto.grid(column=1, row=6, padx=5, pady=5, sticky='w')

        entry_detalles = CTkEntry(frame_formulario_movimientos, placeholder_text="Detalle el destino del dinero", **ESTILO_ENTRYS_LABEL)
        entry_detalles.grid(column=0,row=7,padx=5, pady=5, sticky='w')

        btn_confirmar = CTkButton(frame_formulario_movimientos, text="Confirmar\nmovimiento", **ESTILO_CTKBOTONES,
                                  command=lambda: self.guardar_movimientos(entry_fecha,entry_hora,
                                                                           entry_usuario_logueado,entry_tipo,entry_monto,entry_detalles))
        btn_confirmar.grid(column=1, row=7, padx=5, pady=5,sticky='w') 

       # Crear un Frame para contener el Treeview y el Scrollbar
        tree_frame_movimientos = tk.Frame(self.movimientos_caja,background=COLOR_CUERPO_PRINCIPAL)
        tree_frame_movimientos.pack(padx=5, pady=5)  # Ajuste de columnspan a 2

        # Configurar el estilo
        style = ttk.Style()
        style.configure("mystyle.Treeview", highlightthickness=0, bd=0, font=("OCR A Extended", 10))
        style.configure("mystyle.Treeview.Heading", font=("OCR A Extended", 9, "bold"))
        style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])

        # Crear un Treeview en el frame con el estilo configurado
        movimientos_treeview = ttk.Treeview(tree_frame_movimientos, style="mystyle.Treeview")
        movimientos_treeview.grid(row=0, column=0, sticky="nsew")

        # Configurar las columnas del Treeview
        movimientos_treeview["columns"] = ("ID", "Tipo", "Fecha", "Hora", "Monto", "Usuario", "Detalles")
        movimientos_treeview.heading("#0", text="ID", anchor="center")
        movimientos_treeview.heading("ID", text="ID")
        movimientos_treeview.heading("Tipo", text="Tipo")
        movimientos_treeview.heading("Fecha", text="Fecha")
        movimientos_treeview.heading("Hora", text="Hora")
        movimientos_treeview.heading("Monto", text="Monto")
        movimientos_treeview.heading("Usuario", text="Usuario")
        movimientos_treeview.heading("Detalles", text="Detalles")

        # Ajustar el ancho de las columnas
        movimientos_treeview.column("#0", width=0, stretch=tk.NO)
        movimientos_treeview.column("ID", width=0, stretch=tk.NO)
        movimientos_treeview.column("Tipo", width=100, anchor="center")
        movimientos_treeview.column("Fecha", width=100, anchor="center")
        movimientos_treeview.column("Hora", width=100, anchor="center")
        movimientos_treeview.column("Monto", width=100, anchor="center")
        movimientos_treeview.column("Usuario", width=100, anchor="center")
        movimientos_treeview.column("Detalles", width=150, anchor="center")
        self.ver_movimientos_caja(movimientos_treeview)

        boton_eliminar_movimientos = CTkButton(tree_frame_movimientos,text="Eliminar movimiento",**ESTILO_CTKBOTONES,
                                               command=lambda:self.eliminar_movimiento_caja(movimientos_treeview))
        boton_eliminar_movimientos.grid(column=0,row=1,padx=10,pady=10)

    def guardar_movimientos(self, entry_fecha, entry_hora, entry_usuario_logueado, entry_tipo, entry_monto, entry_detalles):
        try:
            conexion = sqlite3.connect('farmacia.db')
            cursor = conexion.cursor()

            # Obtener la fecha actual
            fecha_actual = datetime.now().strftime('%Y-%m-%d')

            # Verificar si hay una jornada en curso para la fecha actual
            cursor.execute("SELECT * FROM operaciones_caja WHERE fecha_inicio = ? AND estado = ?", (fecha_actual, "Jornada en curso"))
            jornada_en_curso = cursor.fetchone()

            if not jornada_en_curso:
                messagebox.showwarning("Aviso", "No hay una jornada en curso para el día de hoy. Por favor, inicie operaciones.")
                return

            id_movimiento = self.generar_codigo_aleatorio()
            fecha = entry_fecha.get()
            hora = entry_hora.get()
            usuario = entry_usuario_logueado.get()
            tipo = entry_tipo.get()
            monto = float(entry_monto.get())
            detalles = entry_detalles.get()

            cursor.execute('''INSERT INTO movimientos_caja
                            (id, tipo, fecha, hora, monto, usuario, detalles)
                            VALUES (?, ?, ?, ?, ?, ?, ?)''', 
                            (id_movimiento, tipo, fecha, hora, monto, usuario, detalles))
            conexion.commit()

            messagebox.showinfo("Operación exitosa", "El movimiento se ha guardado correctamente.")

        except sqlite3.Error as e:
            if conexion:
                conexion.rollback()
            messagebox.showerror("Error", f"No se pudo guardar el movimiento. Error: {str(e)}")

        finally:
            if conexion:
                conexion.close()
    
    def ver_movimientos_caja(self, movimientos_treeview):
        try:
            conexion = sqlite3.connect('farmacia.db')
            cursor = conexion.cursor()

            # Obtener la fecha actual en formato YYYY-MM-DD
            fecha_actual = datetime.now().strftime('%Y-%m-%d')

            # Obtener los movimientos de caja del día en curso, excluyendo la columna 'id'
            cursor.execute('''SELECT id,tipo, fecha, hora, monto, usuario, detalles 
                            FROM movimientos_caja 
                            WHERE fecha = ?''', (fecha_actual,))
            movimientos = cursor.fetchall()

            # Limpiar el Treeview
            for record in movimientos_treeview.get_children():
                movimientos_treeview.delete(record)

            # Insertar los movimientos en el Treeview
            for movimiento in movimientos:
                movimientos_treeview.insert('', 'end', values=movimiento)

            conexion.commit()

        except sqlite3.Error as e:
            if conexion:
                conexion.rollback()
            messagebox.showerror("Error", f"No se pudo obtener los movimientos. Error: {str(e)}")

        finally:
            if conexion:
                conexion.close()  

    def eliminar_movimiento_caja(self, movimientos_treeview):
        # Obtener el item seleccionado en el Treeview
        seleccion = movimientos_treeview.selection()

        if not seleccion:
            messagebox.showwarning("Advertencia", "Por favor selecciona un movimiento para eliminar.")
            return

        # Obtener el ID del movimiento seleccionado
        id_movimiento = movimientos_treeview.item(seleccion, 'values')[0]
        print(id_movimiento)

        try:
            conexion = sqlite3.connect('farmacia.db')
            cursor = conexion.cursor()

            # Eliminar el movimiento de la base de datos
            cursor.execute("DELETE FROM movimientos_caja WHERE id = ?", (id_movimiento,))
            conexion.commit()

            # Eliminar el movimiento del Treeview
            movimientos_treeview.delete(seleccion)

            messagebox.showinfo("Éxito", "El movimiento se ha eliminado correctamente.")

        except sqlite3.Error as e:
            if conexion:
                conexion.rollback()
            messagebox.showerror("Error", f"No se pudo eliminar el movimiento. Error: {str(e)}")

        finally:
            if conexion:
                conexion.close()

    

                           






#------------------FUNCIONES PARA EL HISTORIAL DE VENTAS ----------------------------------------------------
            
    def historial_ventas_calendario(self):
        # Limpiar cualquier widget existente en el cuerpo principal
        for widget in self.cuerpo_principal.winfo_children():
            widget.destroy()
        
        frame_calendario = tk.Frame(self.cuerpo_principal, background=COLOR_CUERPO_PRINCIPAL)
        frame_calendario.pack(fill="both", padx=10, pady=10, expand=True)

        cal = Calendar(frame_calendario, font=("OCR A Extended", 12), selectmode='day', locale='es_ES', disabledforeground="#858585",
                    cursor="hand2", background="#2A3138",
                    selectbackground="#3B8ED0")
        cal.pack(fill="both", expand=True, padx=10, pady=10)
        cal.bind("<<CalendarSelected>>", lambda event: self.on_date_click(event, treeview_historial_ventas, cal, total_diario_ventas))

        # Crear el estilo para el Treeview
        style = ttk.Style()
        style.configure("mystyle.Treeview", highlightthickness=0, bd=0, font=('OCR A Extended', 9))
        style.configure("mystyle.Treeview.Heading", font=('OCR A Extended', 10, 'bold'))
        
        treeview_historial_ventas = ttk.Treeview(frame_calendario, columns=("ID", "ID Producto","Categoria producto", "Producto", "ID Cliente",
                                                                            "Cliente","Cantidad", "Precio Unitario", "Total", 
                                                                            "Método de Pago", "Fecha", "Hora"), style="mystyle.Treeview", show="headings")
        treeview_historial_ventas.pack(fill="both", expand=True, padx=10, pady=10)
        
        treeview_historial_ventas.column("ID", width=0, stretch=FALSE)
        treeview_historial_ventas.column("ID Producto", width=0, stretch=FALSE)
        treeview_historial_ventas.column("ID Cliente", width=0, stretch=FALSE)
        
        treeview_historial_ventas.column("Producto", width=70)
        treeview_historial_ventas.column("Categoria producto", width=70)
        treeview_historial_ventas.column("Cliente", width=70)
        treeview_historial_ventas.column("Cantidad", width=10, anchor="center")
        treeview_historial_ventas.column("Precio Unitario", width=10, anchor="center")
        treeview_historial_ventas.column("Total", width=10, anchor="center")
        treeview_historial_ventas.column("Método de Pago", width=70, anchor="center")
        treeview_historial_ventas.column("Fecha", width=70)
        treeview_historial_ventas.column("Hora", width=70)

        # Configurar encabezados del Treeview
        treeview_historial_ventas.heading("ID", text="ID")
        treeview_historial_ventas.heading("ID Producto", text="ID Producto")
        treeview_historial_ventas.heading("Categoria producto", text="Categoria")
        treeview_historial_ventas.heading("Producto", text="Producto")
        treeview_historial_ventas.heading("ID Cliente", text="ID Cliente")
        treeview_historial_ventas.heading("Cliente", text="Cliente")
        treeview_historial_ventas.heading("Cantidad", text="Cantidad")
        treeview_historial_ventas.heading("Precio Unitario", text="PvP")
        treeview_historial_ventas.heading("Total", text="Total")
        treeview_historial_ventas.heading("Método de Pago", text="Método Pago")
        treeview_historial_ventas.heading("Fecha", text="Fecha")
        treeview_historial_ventas.heading("Hora", text="Hora")

        total_diario_ventas = CTkLabel(frame_calendario, text="TOTAL VENTAS DIARIAS: $", width=70, height=50, font=("OCR A Extended", 25),
                        anchor="center", bg_color=COLOR_CUERPO_PRINCIPAL, corner_radius=32)
        total_diario_ventas.pack(side="right", padx=10, pady=10)

        eliminar_registro = CTkButton(frame_calendario, text="Eliminar registro", width=70, height=50, text_color='black',
                            font=("OCR A Extended", 12),
                            command=lambda: self.borrar_registro(treeview_historial_ventas, total_diario_ventas))
        eliminar_registro.pack(side="left", padx=10, pady=10)

        # Frame para el gráfico de barras
        self.frame_grafico = ttk.Frame(self.cuerpo_principal, width=300)  # Ancho ajustable según tus necesidades
        self.frame_grafico.pack(side="right", fill="both", expand=True, padx=20, pady=20)

        today = datetime.today().date()
        self.load_sales(today, today, treeview_historial_ventas, total_diario_ventas)

    def borrar_registro(self, treeview_historial_ventas, total_diario_ventas):
        selected_item = treeview_historial_ventas.selection()

        if selected_item:
            # Obtener el ID de la venta seleccionada
            venta_id = treeview_historial_ventas.item(selected_item, "values")[0]

            # Conectar a la base de datos
            conn = sqlite3.connect('farmacia.db')
            cursor = conn.cursor()

            try:
                # Obtener las ventas eliminadas de la base de datos
                cursor.execute("SELECT id_producto, cantidad FROM ventas WHERE id = ?", (venta_id,))
                ventas_eliminadas = cursor.fetchall()

                # Eliminar la venta de la base de datos
                cursor.execute("DELETE FROM ventas WHERE id = ?", (venta_id,))
                conn.commit()

                # Actualizar existencias de productos
                for venta in ventas_eliminadas:
                    id_producto, cantidad_vendida = venta
                    cursor.execute("UPDATE productos SET existencias = existencias + ? WHERE id = ?", (cantidad_vendida, id_producto))
                    conn.commit()

                # Cerrar la conexión a la base de datos
                conn.close()

                # Eliminar la fila seleccionada del treeview
                treeview_historial_ventas.delete(selected_item)

                # Obtener la fecha actual
                current_date = datetime.now().date()
                print(current_date)

                # Actualizar el total de ventas diarias
                self.load_sales(current_date, current_date, treeview_historial_ventas, total_diario_ventas)

            except sqlite3.Error as e:
                # Manejar cualquier error en la operación de la base de datos
                messagebox.showerror("Error", f"No se pudo eliminar la venta y actualizar el inventario.\nError: {e}")

                # Cerrar la conexión en caso de error
                conn.close()
        else:
            messagebox.showwarning("Seleccione una Venta", "Por favor, seleccione una venta para eliminar.")

    def on_date_click(self, event, treeview_historial_ventas, cal, total_diario_ventas):
        selected_date = cal.get_date()
        try:
            # Intentar convertir la fecha al formato esperado
            start_date = datetime.strptime(selected_date, "%d/%m/%Y").date()
        except ValueError:
            # Si falla, intentar con un formato alternativo
            start_date = datetime.strptime(selected_date, "%d/%m/%y").date()
        
        end_date = start_date + timedelta(days=1)  # Incrementar la fecha en un día para obtener el rango completo del día seleccionado
        self.load_sales(start_date, end_date, treeview_historial_ventas, total_diario_ventas)

    def load_sales(self, start_date, end_date, treeview_historial_ventas, total_diario_ventas):
        # Limpiar cualquier venta existente en el Treeview
        for item in treeview_historial_ventas.get_children():
            treeview_historial_ventas.delete(item)

        # Realizar consulta en la base de datos para obtener las ventas del día seleccionado
        conn = sqlite3.connect('farmacia.db')
        c = conn.cursor()
        c.execute("SELECT * FROM ventas WHERE fecha >= ? AND fecha < ?", (start_date, end_date))
        sales = c.fetchall()
        
        # Calcular el total de las ventas diarias
        total_ventas_diarias = sum(sale[7] for sale in sales)  # Se asume que el total de la venta está en la posición 7

        # Actualizar el texto del label con el total de las ventas diarias
        total_diario_ventas.configure(text=f"TOTAL VENTAS DIARIAS: ${total_ventas_diarias:.2f}")
        
        conn.close()

        # Mostrar las ventas en el Treeview
        for sale in sales:
            treeview_historial_ventas.insert('', 'end', values=sale)
    
    def generar_grafico_barras(self, data):
        productos = [sale[1] for sale in data]
        cantidades = [sale[2] for sale in data]

        # Limpiar el frame del gráfico
        for widget in self.frame_grafico.winfo_children():
            widget.destroy()

        fig, ax = plt.subplots()
        ax.bar(productos, cantidades)
        ax.set_xlabel('Productos')
        ax.set_ylabel('Cantidad')
        ax.set_title('Productos Más Vendidos')

        canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

       # canvas.get_tk_widget().configure(yscrollcommand=None, xscrollcommand=None)
    



        




#----------clase parpadeo etiqueta--------------------------------------------------------

class ParpadeoEtiqueta(tk.Label):
    def __init__(self, parent, text, *args, **kwargs):
        tk.Label.__init__(self, parent, text=text,background=COLOR_CUERPO_PRINCIPAL, *args, **kwargs)
        self.parpadear()

    def parpadear(self):
        current_color = self.cget("foreground")
        new_color = "red" if current_color == "black" else "black"
        self.config(foreground=new_color)
        self.after(500, self.parpadear)  # Cambia el valor de 500 para ajustar la velocidad del parpadeo
""""

#------------------------FUNCIONES PARA CLIENTE-------------------------------

    def mostrar_clientes(self):
        # Definir el estilo general para los botones
        ESTILO_CTKBOTONES = {
            'width': 70,
            'height': 70,
            'text_color': 'black',
            'font': ("OCR A Extended", 13,"bold")
        }
        # Función para mostrar la lista de clientes
        for widget in self.cuerpo_principal.winfo_children():
            widget.destroy()
        

        frame_tree_clientes = tk.Frame(self.cuerpo_principal,bg=COLOR_CUERPO_PRINCIPAL)
        frame_tree_clientes.pack(pady=10, padx=10, side="right", fill="both", expand=True)
        
        # Crear el estilo para el Treeview
        style = ttk.Style()
        style.configure("mystyle.Treeview", highlightthickness=0, bd=0, font=('OCR A Extended', 10))
        style.configure("mystyle.Treeview.Heading", font=('OCR A Extended', 11, 'bold'))
        style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])

        # Aplicar el estilo al Treeview
        tabla_cliente = ttk.Treeview(frame_tree_clientes, columns=("Cédula", "Nombre", "Teléfono", "Dirección", "Email"), displaycolumns=("Cédula", "Nombre", "Teléfono", "Dirección", "Email"), style="mystyle.Treeview")
        tabla_cliente.heading("#0", text="", anchor="center")  # Eliminar encabezado de índice
        tabla_cliente.heading("Cédula", text="CEDULA")
        tabla_cliente.heading("Nombre", text="NOMBRE")
        tabla_cliente.heading("Teléfono", text="TELEFONO")
        tabla_cliente.heading("Dirección", text="DIRECCION")
        tabla_cliente.heading("Email", text="EMAIL")
        tabla_cliente.column("#0", width=0, stretch=tk.NO)  # Hacer que la columna de índice no sea visible
        tabla_cliente.column("Cédula", width=70)
        tabla_cliente.column("Nombre", width=110)
        tabla_cliente.column("Teléfono", width=70)
        tabla_cliente.column("Dirección", width=120)
        tabla_cliente.column("Email", width=150)
        tabla_cliente.pack(side="top", expand=True, fill="both")

        # Botón Editar
        btn_editar = CTkButton(frame_tree_clientes, text="Editar", **ESTILO_CTKBOTONES)
        btn_editar.pack(side="left", padx=5, pady=5)

        # Botón Eliminar
        btn_eliminar = CTkButton(frame_tree_clientes, text="Eliminar", **ESTILO_CTKBOTONES)
        btn_eliminar.pack(side="left", padx=5, pady=5)

        # Formulario de ingreso de cliente
        frame_formulario_cliente = tk.Frame(self.cuerpo_principal,bg=COLOR_CUERPO_PRINCIPAL)
        frame_formulario_cliente.pack(pady=10, side="left")

        lbl_titulo = CTkLabel(frame_formulario_cliente, text="Ingreso de Clientes", font=("OCR A Extended", 12, "bold"))
        lbl_titulo.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        lbl_id = CTkLabel(frame_formulario_cliente, text="id:",font=("OCR A Extended", 10, "bold"))
        lbl_id.grid(row=2, column=0, padx=5, pady=5, sticky="w")
        entry_id = CTkEntry(frame_formulario_cliente,font=("OCR A Extended", 10, "bold"))
        entry_id.grid(row=2, column=1, padx=5, pady=5)
    
        lbl_cedula = CTkLabel(frame_formulario_cliente, text="Cédula:",font=("OCR A Extended", 10, "bold"))
        lbl_cedula.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        entry_cedula = CTkEntry(frame_formulario_cliente,font=("OCR A Extended", 10, "bold"))
        entry_cedula.grid(row=3, column=1, padx=5, pady=5)

        lbl_nombre = CTkLabel(frame_formulario_cliente, text="Nombre:",font=("OCR A Extended", 10, "bold"))
        lbl_nombre.grid(row=4, column=0, padx=5, pady=5, sticky="w")
        entry_nombre = CTkEntry(frame_formulario_cliente,font=("OCR A Extended", 10, "bold"))
        entry_nombre.grid(row=4, column=1, padx=5, pady=5)

        lbl_telefono = CTkLabel(frame_formulario_cliente, text="Teléfono:",font=("OCR A Extended", 10, "bold"))
        lbl_telefono.grid(row=5, column=0, padx=5, pady=5, sticky="w")
        entry_telefono = CTkEntry(frame_formulario_cliente,font=("OCR A Extended", 10, "bold"))
        entry_telefono.grid(row=5, column=1, padx=5, pady=5)

        lbl_direccion = CTkLabel(frame_formulario_cliente, text="Dirección:",font=("OCR A Extended", 10, "bold"))
        lbl_direccion.grid(row=6, column=0, padx=5, pady=5, sticky="w")
        entry_direccion = CTkEntry(frame_formulario_cliente,font=("OCR A Extended", 10, "bold"))
        entry_direccion.grid(row=6, column=1, padx=5, pady=5)

        lbl_email = CTkLabel(frame_formulario_cliente, text="Email:",font=("OCR A Extended", 10, "bold"))
        lbl_email.grid(row=7, column=0, padx=5, pady=5, sticky="w")
        entry_email = CTkEntry(frame_formulario_cliente,font=("OCR A Extended", 10, "bold"))
        entry_email.grid(row=7, column=1, padx=5, pady=5)

        btn_guardar = CTkButton(frame_formulario_cliente, text="Guardar",**ESTILO_CTKBOTONES)
        btn_guardar.grid(row=8, column=1, padx=5, pady=5, sticky="ns")
        
#--------------------------FUNCIONES PARA DATOS DEL NEGOCIO--------------------------------
    def datos_negocio(self):

        # Definir el estilo general para los botones
        ESTILO_CTKBOTONES_DATOS_NEGO = {
            'width': 50,
            'height': 50,
            'text_color': 'black',
            'font': ("OCR A Extended", 13,"bold")
        }

        ESTILO_TITULO_LABEL_DATOS_NEGO = {
            'text_color': 'black',
            'font': ("OCR A Extended", 15, "bold"),  
        }

        ESTILO_ENTRYS_LABEL_DATOS_NEGO = {
            'text_color': 'black',
            'font': ("OCR A Extended", 13),
            'width': 200      
        }

        global conexion, cursor
        conexion = sqlite3.connect('farmacia.db')
        cursor = conexion.cursor()
        for widget in self.cuerpo_principal.winfo_children():
            widget.destroy()

        def generar_clave():
            caracteres = string.ascii_letters + string.digits
            return ''.join(random.choice(caracteres) for _ in range(8))

        def insertar_datos(conexion):
            try:
                # Obtener los valores de los campos de entrada
                datos = (nombre_negocio.get(), ruc.get(), direccion.get(), telefono.get(), email.get())

                # Verificar si hay datos existentes
                cursor.execute('SELECT COUNT(*) FROM datos_negocio')
                count = cursor.fetchone()[0]

                if count == 0:
                    # Si no hay registros, insertar uno nuevo
                    cursor.execute('INSERT INTO datos_negocio (nombre_negocio, ruc, direccion, telefono, email) VALUES (?, ?, ?, ?, ?)', datos)
                else:
                    # Si ya hay un registro, actualizarlo
                    cursor.execute('UPDATE datos_negocio SET nombre_negocio=?, ruc=?, direccion=?, telefono=?, email=?', datos)

                conexion.commit()
                mostrar_datos_actuales()
                messagebox.showinfo("Éxito", "Los datos se han guardado correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudieron guardar los datos. Error: {str(e)}")



                # Función para llenar los entrys con los datos actuales del negocio
        def llenar_entrys():
            cursor.execute('SELECT * FROM datos_negocio')
            datos = cursor.fetchone()
            if datos:
                # Obtener el id del registro actual para su posterior actualización
                id_registro = datos[0]

                # Actualizar los campos de entrada con los datos recuperados de la base de datos
                nombre_negocio.delete(0, 'end')
                nombre_negocio.insert(0, datos[1])
                ruc.delete(0, 'end')
                ruc.insert(0, datos[2])
                direccion.delete(0, 'end')
                direccion.insert(0, datos[3])
                telefono.delete(0, 'end')
                telefono.insert(0, datos[4])
                email.delete(0, 'end')
                email.insert(0, datos[5])

                # Retornar el id del registro actual para su posterior uso en la actualización
                return id_registro
            else:
                messagebox.showerror("Error", "No hay datos para mostrar.")
                return None


        # Función para mostrar los datos actuales del negocio sin llenar los entrys
        def mostrar_datos_actuales():
            cursor.execute('SELECT * FROM datos_negocio')
            datos = cursor.fetchone()
            if datos:
                lbl_nombre.configure(text=datos[1])
                lbl_ruc.configure(text=datos[2])
                lbl_direccion.configure(text=datos[3])
                lbl_telefono.configure(text=datos[4])
                lbl_email.configure(text=datos[5])

        # Crear Frames
        frame_formulario = tk.Frame(self.cuerpo_principal,bg=COLOR_CUERPO_PRINCIPAL)
        frame_formulario.grid(row=0,columnspan=2,column=0, padx=10, pady=10)

        frame_datos_actuales = tk.Frame(self.cuerpo_principal,bg=COLOR_CUERPO_PRINCIPAL)
        frame_datos_actuales.grid(row=1, column=0, padx=10, pady=10)

        # Crear etiquetas y entradas para el formulario
        lbl_titulo_formulario_nego = CTkLabel(frame_formulario, text="I N G R E S O   D E   D A T O S",**ESTILO_TITULO_LABEL_DATOS_NEGO )
        lbl_titulo_formulario_nego.grid(row=0, columnspan=2, padx=5, pady=5, sticky="w")

        CTkLabel(frame_formulario, text="Nombre del Negocio:",**ESTILO_ENTRYS_LABEL_DATOS_NEGO).grid(row=1, column=0, padx=5, pady=5, sticky="e")
        nombre_negocio = CTkEntry(frame_formulario,**ESTILO_ENTRYS_LABEL_DATOS_NEGO)
        nombre_negocio.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        CTkLabel(frame_formulario, text="RUC:",**ESTILO_ENTRYS_LABEL_DATOS_NEGO).grid(row=2, column=0, padx=5, pady=5, sticky="e")
        ruc = CTkEntry(frame_formulario,**ESTILO_ENTRYS_LABEL_DATOS_NEGO)
        ruc.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        CTkLabel(frame_formulario, text="Dirección:",**ESTILO_ENTRYS_LABEL_DATOS_NEGO).grid(row=3, column=0, padx=5, pady=5, sticky="e")
        direccion = CTkEntry(frame_formulario,**ESTILO_ENTRYS_LABEL_DATOS_NEGO)
        direccion.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        CTkLabel(frame_formulario, text="Teléfono:",**ESTILO_ENTRYS_LABEL_DATOS_NEGO).grid(row=4, column=0, padx=5, pady=5, sticky="e")
        telefono = CTkEntry(frame_formulario,**ESTILO_ENTRYS_LABEL_DATOS_NEGO)
        telefono.grid(row=4, column=1, padx=5, pady=5, sticky="w")

        CTkLabel(frame_formulario, text="Email:",**ESTILO_ENTRYS_LABEL_DATOS_NEGO).grid(row=5, column=0, padx=5, pady=5, sticky="e")
        email = CTkEntry(frame_formulario,**ESTILO_ENTRYS_LABEL_DATOS_NEGO)
        email.grid(row=5, column=1, padx=5, pady=5, sticky="w")
        
        # Botón para guardar los datos
        btn_guardar = CTkButton(frame_formulario, text="Guardar Datos", **ESTILO_CTKBOTONES_DATOS_NEGO, command=lambda: insertar_datos(conexion))
        btn_guardar.grid(row=6, columnspan=2, padx=5, pady=5)
        
        # Crear etiquetas para mostrar los datos actuales del negocio
        lbl_nombre = CTkLabel(frame_datos_actuales, text="", **ESTILO_ENTRYS_LABEL_DATOS_NEGO)
        lbl_nombre.grid(row=0, padx=5, pady=5, sticky="nsew")

        lbl_ruc = CTkLabel(frame_datos_actuales, text="", **ESTILO_ENTRYS_LABEL_DATOS_NEGO)
        lbl_ruc.grid(row=1, padx=5, pady=5, sticky="nsew")

        lbl_direccion = CTkLabel(frame_datos_actuales, text="", **ESTILO_ENTRYS_LABEL_DATOS_NEGO)
        lbl_direccion.grid(row=2, padx=5, pady=5, sticky="nsew")

        lbl_telefono = CTkLabel(frame_datos_actuales, text="", **ESTILO_ENTRYS_LABEL_DATOS_NEGO)
        lbl_telefono.grid(row=3, padx=5, pady=5, sticky="nsew")

        lbl_email = CTkLabel(frame_datos_actuales, text="", **ESTILO_ENTRYS_LABEL_DATOS_NEGO)
        lbl_email.grid(row=4, padx=5, pady=5, sticky="nsew")

        # Crear botón "Editar" para llenar los entrys con los datos actuales
        btn_editar = CTkButton(frame_datos_actuales, text="Editar Datos", **ESTILO_CTKBOTONES_DATOS_NEGO, command=llenar_entrys)
        btn_editar.grid(row=5, padx=5, pady=5, sticky="nsew")

        mostrar_datos_actuales()
   

"""




# Conectar a la base de datos (o crearla si no existe)
conn = sqlite3.connect('farmacia.db')
cursor = conn.cursor()

# Crear la tabla de productos
cursor.execute('''CREATE TABLE IF NOT EXISTS productos (
                    id TEXT PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    descripcion TEXT,
                    categoria TEXT,
                    precio_venta REAL,
                    precio_compra REAL,
                    existencias INTEGER,
                    stock_minimo INTEGER,
                    codigo_barras TEXT        
                )''')

cursor.execute('''CREATE TABLE IF NOT EXISTS usuarios (
                    id TEXT PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    password TEXT,
                    rol TEXT,
                    estado REAL
                )''')


# Crear la tabla de ventas
cursor.execute('''CREATE TABLE IF NOT EXISTS ventas (
                    id TEXT PRIMARY KEY,
                    id_producto TEXT,
                    categoria_producto TEXT,
                    producto TEXT,
                    id_cliente TEXT,
                    cliente TEXT,
                    cantidad INTEGER,
                    precio_unitario REAL,
                    total REAL,
                    metodo_pago TEXT,
                    fecha DATE,
                    hora TIME,
                    FOREIGN KEY (id_producto) REFERENCES productos (id)
                )''')



# Crear la tabla de proveedores
cursor.execute('''CREATE TABLE IF NOT EXISTS proveedores (
                    id TEXT PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    contacto TEXT
                )''')

# Crear la tabla de clientes
cursor.execute('''CREATE TABLE IF NOT EXISTS clientes (
                    id TEXT PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    cedula TEXT,
                    direccion TEXT,
                    telefono TEXT,
                    Email TEXT
                )''')

# Crear la tabla de datos negocio
cursor.execute('''CREATE TABLE IF NOT EXISTS datos_negocio (
                    id TEXT PRIMARY KEY,
                    nombre_negocio TEXT NOT NULL,
                    ruc TEXT,
                    direccion TEXT,
                    telefono TEXT,
                    Email TEXT
                )''')

cursor.execute('''CREATE TABLE IF NOT EXISTS operaciones_caja (
                    id TEXT PRIMARY KEY,
                    fecha_inicio DATE,
                    hora_inicio TIME,
                    fecha_final DATE,
                    hora_final TIME,
                    usuario TEXT,
                    estado TEXT,
                    valor_inicial REAL,
                    valor_cierre REAL,
                    comentarios TEXT,
                    detalles_movimientos TEXT,
                    detalles_transacciones TEXT
                )''')

# Crear la tabla de movimientos caja
cursor.execute('''CREATE TABLE IF NOT EXISTS movimientos_caja (
                    id TEXT PRIMARY KEY,
                    tipo TEXT NOT NULL,
                    fecha DATE,
                    hora TIME,
                    monto REAL,
                    usuario TEXT,
                    detalles TEXT
                )''')


# Guardar cambios y cerrar la conexión
conn.commit()
conn.close()